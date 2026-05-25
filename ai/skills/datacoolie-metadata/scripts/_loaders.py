"""Shared file-loading helpers for datacoolie-metadata skill scripts.

Provides consistent JSON / YAML / Excel loading for all scripts.
Named _loaders (not _io) to avoid collision with Python's built-in _io module.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import yaml


# -- Excel native format constants (mirrors file_provider.py) ---------------

_EXCEL_LIST_COLS: frozenset = frozenset({
    "source_watermark_columns",
    "destination_merge_keys",
})

# Columns that accept JSON arrays OR comma-separated strings, but items
# must be PartitionColumn objects ({"column": "..."}) per schema.
_EXCEL_PARTITION_COLS: frozenset = frozenset({
    "destination_partition_columns",
})

_EXCEL_JSON_COLS: frozenset = frozenset({
    "configure",
    "source_configure",
    "destination_configure",
    "transform_deduplicate_columns",
    "transform_latest_data_columns",
    "transform_additional_columns",
    "transform_schema_hints",
    "transform_configure",
})


def _cast(value: Any) -> Any:
    """Normalise cell values: strip strings, return None for empty."""
    if isinstance(value, str):
        value = value.strip()
        return value if value != "" else None
    return value


def _safe_bool(value: Any) -> bool:
    """Coerce a cell to bool; returns False for unrecognised values."""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        return value.strip().lower() in ("true", "1", "yes")
    return False


def _json_cell(value: Any) -> Any:
    """Parse a JSON object/array cell; return None if blank."""
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    if s.startswith(("{", "[")):
        return json.loads(s)
    return None


def _ensure_list(value: Any) -> list:
    """Convert comma-separated string or single value to list."""
    if value is None:
        return []
    if isinstance(value, list):
        return value
    s = str(value).strip()
    if not s:
        return []
    return [item.strip() for item in s.split(",") if item.strip()]


def _parse_partition_columns(value: Any) -> list:
    """Parse partition_columns cell: JSON array of objects OR comma-separated names.

    Returns list of PartitionColumn dicts: [{"column": "name"}, ...] or
    the parsed JSON array as-is if it already contains objects.
    """
    if value is None:
        return []
    if isinstance(value, list):
        return value
    s = str(value).strip()
    if not s:
        return []
    # JSON array → parse directly (objects with column/expression)
    if s.startswith("["):
        try:
            return json.loads(s)
        except json.JSONDecodeError:
            pass
    # Comma-separated column names → wrap as PartitionColumn objects
    names = [item.strip() for item in s.split(",") if item.strip()]
    return [{"column": name} for name in names]


def _excel_sheet_rows(wb: Any, sheet_name: str) -> list[dict[str, Any]]:
    """Return a list of row-dicts for sheet_name, or [] if absent."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    rows: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        rows.append({headers[i]: row[i] for i in range(len(headers))})
    return rows


def _parse_excel_connections(wb: Any) -> list[dict[str, Any]]:
    """Parse connections sheet (native datacoolie xlsx format)."""
    connections: list[dict[str, Any]] = []
    for row in _excel_sheet_rows(wb, "connections"):
        conn: dict[str, Any] = {}
        cfg: dict[str, Any] = {}
        for key, raw_val in row.items():
            val = _cast(raw_val)
            if key == "configure":
                parsed = _json_cell(raw_val)
                if isinstance(parsed, dict):
                    cfg.update(parsed)
            elif key.startswith("configure_"):
                sub = key[len("configure_"):]
                if val is not None:
                    cfg[sub] = val
            elif key == "secrets_ref":
                parsed = _json_cell(raw_val)
                if parsed is not None:
                    conn[key] = parsed
            elif key == "is_active":
                if val is not None:
                    conn[key] = _safe_bool(val)
            elif val is not None:
                conn[key] = val
        if cfg:
            conn["configure"] = cfg
        if conn:
            connections.append(conn)
    return connections


def _parse_excel_dataflows(wb: Any) -> list[dict[str, Any]]:
    """Parse dataflows sheet (native datacoolie xlsx format)."""
    dataflows: list[dict[str, Any]] = []
    for row in _excel_sheet_rows(wb, "dataflows"):
        df_dict: dict[str, Any] = {}
        src: dict[str, Any] = {}
        dest: dict[str, Any] = {}
        transform: dict[str, Any] = {}
        transform_base: dict[str, Any] = {}
        for key, raw_val in row.items():
            if key == "transform":
                parsed = _json_cell(raw_val)
                if isinstance(parsed, dict):
                    transform_base = parsed
                continue
            if key == "is_active":
                parsed_val = _cast(raw_val)
                if parsed_val is not None:
                    df_dict["is_active"] = _safe_bool(parsed_val)
                continue
            if key.startswith("source_"):
                target, sub = src, key[len("source_"):]
            elif key.startswith("destination_"):
                target, sub = dest, key[len("destination_"):]
            elif key.startswith("transform_"):
                target, sub = transform, key[len("transform_"):]
            else:
                val = _json_cell(raw_val) if key in _EXCEL_JSON_COLS else _cast(raw_val)
                if val is not None:
                    df_dict[key] = val
                continue
            if key in _EXCEL_LIST_COLS:
                parsed_list = _ensure_list(raw_val)
                if parsed_list:
                    target[sub] = parsed_list
            elif key in _EXCEL_PARTITION_COLS:
                # JSON array of objects OR comma-separated column names
                parsed_pc = _parse_partition_columns(raw_val)
                if parsed_pc:
                    target[sub] = parsed_pc
            elif key in _EXCEL_JSON_COLS:
                parsed_json = _json_cell(raw_val)
                if parsed_json is not None:
                    target[sub] = parsed_json
            else:
                val = _cast(raw_val)
                if val is not None:
                    target[sub] = val
        if src:
            df_dict["source"] = src
        if dest:
            df_dict["destination"] = dest
        merged_transform = {**transform_base, **transform}
        if merged_transform:
            df_dict["transform"] = merged_transform
        if df_dict:
            dataflows.append(df_dict)
    return dataflows


def _parse_excel_schema_hints(wb: Any) -> list[dict[str, Any]]:
    """Parse schema_hints sheet (native datacoolie xlsx format)."""
    hints_map: dict[tuple, dict[str, Any]] = {}
    for row in _excel_sheet_rows(wb, "schema_hints"):
        conn_name = _cast(row.get("connection_name"))
        table_name = _cast(row.get("table_name"))
        schema_name = _cast(row.get("schema_name"))
        if not conn_name or not table_name:
            continue
        group_key = (conn_name, table_name, schema_name)
        if group_key not in hints_map:
            hints_map[group_key] = {
                "connection_name": conn_name,
                "table_name": table_name,
                "hints": [],
            }
            if schema_name is not None:
                hints_map[group_key]["schema_name"] = schema_name
        hint: dict[str, Any] = {}
        for field in ("column_name", "data_type", "format"):
            val = _cast(row.get(field))
            if val is not None:
                hint[field] = val
        for field in ("precision", "scale"):
            raw = row.get(field)
            if raw is not None:
                try:
                    hint[field] = int(raw)
                except (ValueError, TypeError):
                    pass
        if hint:
            hints_map[group_key]["hints"].append(hint)
    return list(hints_map.values())


def load_excel(file_path: Path) -> dict:
    """Load an Excel (.xlsx/.xls) metadata file in native datacoolie format.

    Expects sheets: connections, dataflows, schema_hints (all optional).
    Returns the same dict structure as JSON/YAML metadata.

    Raises:
        ImportError: if openpyxl is not installed.
    """
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError(
            "openpyxl is required for Excel metadata files. "
            "Install it with: pip install openpyxl"
        ) from exc
    wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
    try:
        result: dict[str, Any] = {}
        connections = _parse_excel_connections(wb)
        if connections:
            result["connections"] = connections
        dataflows = _parse_excel_dataflows(wb)
        if dataflows:
            result["dataflows"] = dataflows
        schema_hints = _parse_excel_schema_hints(wb)
        if schema_hints:
            result["schema_hints"] = schema_hints
        return result
    finally:
        wb.close()


# -- Public API -------------------------------------------------------------


def load_file(file_path: Path) -> dict | list:
    """Load a JSON, YAML, or Excel file. Returns the parsed content.

    Raises:
        json.JSONDecodeError: on malformed JSON.
        yaml.YAMLError: on malformed YAML.
        ImportError: if openpyxl needed but not installed.
    """
    suffix = file_path.suffix.lower()
    if suffix in (".xlsx", ".xls"):
        return load_excel(file_path)
    with open(file_path, encoding="utf-8") as f:
        if suffix in (".yaml", ".yml"):
            return yaml.safe_load(f) or {}
        return json.load(f)


def load_metadata(file_path: Path) -> dict:
    """Load a metadata JSON, YAML, or Excel file. Always returns a dict.

    Raises:
        json.JSONDecodeError: on malformed JSON.
        yaml.YAMLError: on malformed YAML.
        ImportError: if openpyxl needed but not installed.
        ValueError: if the top-level value is not a JSON object.
    """
    data = load_file(file_path)
    if not isinstance(data, dict):
        raise ValueError(
            f"{file_path}: expected a JSON object at the top level, got {type(data).__name__}"
        )
    return data
