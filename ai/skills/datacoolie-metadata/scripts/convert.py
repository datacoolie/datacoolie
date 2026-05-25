"""Convert DataCoolie metadata between JSON, YAML, and Excel formats.

Usage:
    python convert.py <input_file> --to json|yaml|excel [--output <path>]
    python convert.py metadata.json --to yaml
    python convert.py metadata.json --to excel
    python convert.py metadata.xlsx --to json
    python convert.py metadata.xlsx --to yaml --output out/metadata.yaml
    python convert.py metadata.yaml --to json --output out/metadata.json

Bidirectional:
    JSON ↔ YAML ↔ Excel (all directions supported)

Excel format (native datacoolie format):
    - connections sheet: name, connection_type, format, catalog, database,
        configure (JSON), secrets_ref (JSON), is_active
    - dataflows sheet: flat columns with source_* / destination_* / transform_*
        prefixes matching file_provider.py convention
    - schema_hints sheet: connection_name, table_name, schema_name, column_name,
        data_type, format, precision, scale, is_active

Requires: pip install openpyxl (for Excel)

Exit codes:
    0 = success
    1 = conversion/validation error
    2 = input error
"""

import argparse
import json
import sys
from pathlib import Path
from typing import Any

import yaml

from _loaders import load_file


class OrderedDumper(yaml.SafeDumper):
    """YAML dumper that preserves insertion order and uses readable formatting."""
    pass


def _dict_representer(dumper, data):
    return dumper.represent_mapping(yaml.resolver.BaseResolver.DEFAULT_MAPPING_TAG, data.items())


OrderedDumper.add_representer(dict, _dict_representer)


def to_json(data: dict) -> str:
    """Serialize dict to pretty JSON."""
    return json.dumps(data, indent=2, ensure_ascii=False) + "\n"


def to_yaml(data: dict) -> str:
    """Serialize dict to readable YAML."""
    return yaml.dump(
        data,
        Dumper=OrderedDumper,
        default_flow_style=False,
        allow_unicode=True,
        sort_keys=False,
        width=120,
    )


# ---------------------------------------------------------------------------
# Native Excel format helpers
# ---------------------------------------------------------------------------

# Fixed column order for each sheet — matches local_use_cases.xlsx convention.
_CONNECTION_COLS = [
    "name", "connection_type", "format", "catalog", "database",
    "configure", "secrets_ref", "is_active",
]
_DATAFLOW_COLS = [
    "name", "stage", "group_number", "execution_order", "processing_mode",
    "is_active", "configure",
    "source_connection_name", "source_schema_name", "source_table",
    "source_query", "source_watermark_columns", "source_configure",
    "destination_connection_name", "destination_schema_name", "destination_table",
    "destination_load_type", "destination_merge_keys",
    "destination_partition_columns", "destination_configure",
    "transform",
]
_SCHEMA_HINTS_COLS = [
    "connection_name", "table_name", "schema_name",
    "column_name", "data_type", "format", "precision", "scale", "is_active",
]


def _json_or_empty(value: object) -> str:
    """Serialize a dict/list to a compact JSON string, or empty string for None."""
    if value is None:
        return ""
    return json.dumps(value, ensure_ascii=False)


def _list_to_cell(items: list) -> str:
    """Convert a list to a comma-separated string."""
    return ", ".join(str(i) for i in items) if items else ""


def _partition_cols_to_cell(items: list) -> str:
    """Serialize partition_columns:
    - Simple {column only} items → comma-separated names
    - Any item with expression → JSON array for full fidelity
    """
    if not items:
        return ""
    # If every item is a plain string or a dict with only 'column' key
    if all(
        (isinstance(item, str)) or
        (isinstance(item, dict) and list(item.keys()) == ["column"])
        for item in items
    ):
        names = [item if isinstance(item, str) else item["column"] for item in items]
        return ", ".join(names)
    # Complex: has expressions → serialize as JSON
    return json.dumps(items, ensure_ascii=False)


def _connection_to_row(conn: dict) -> dict:
    """Flatten a connection dict to a native xlsx row."""
    row: dict = {}
    for field in ("name", "connection_type", "format", "catalog", "database"):
        row[field] = conn.get(field, "")
    row["configure"] = _json_or_empty(conn.get("configure")) or ""
    row["secrets_ref"] = _json_or_empty(conn.get("secrets_ref")) or ""
    ia = conn.get("is_active")
    row["is_active"] = "" if ia is None else ia
    return row


def _dataflow_to_row(df: dict) -> dict:
    """Flatten a dataflow dict to a native xlsx row (source_* / destination_* prefixes)."""
    row: dict = {}
    # Top-level scalar fields
    for field in ("name", "stage", "group_number", "execution_order", "processing_mode"):
        row[field] = df.get(field, "")
    ia = df.get("is_active")
    row["is_active"] = "" if ia is None else ia
    row["configure"] = _json_or_empty(df.get("configure")) or ""

    # Source fields
    src = df.get("source", {})
    row["source_connection_name"] = src.get("connection_name", "")
    row["source_schema_name"] = src.get("schema_name", "")
    row["source_table"] = src.get("table", "")
    row["source_query"] = src.get("query", "")
    row["source_watermark_columns"] = _list_to_cell(src.get("watermark_columns", []))
    row["source_configure"] = _json_or_empty(src.get("configure")) or ""

    # Destination fields
    dest = df.get("destination", {})
    row["destination_connection_name"] = dest.get("connection_name", "")
    row["destination_schema_name"] = dest.get("schema_name", "")
    row["destination_table"] = dest.get("table", "")
    row["destination_load_type"] = dest.get("load_type", "")
    row["destination_merge_keys"] = _list_to_cell(dest.get("merge_keys", []))
    row["destination_partition_columns"] = _partition_cols_to_cell(dest.get("partition_columns", []))
    row["destination_configure"] = _json_or_empty(dest.get("configure")) or ""

    # Transform: serialize full transform dict as JSON
    transform = df.get("transform")
    row["transform"] = _json_or_empty(transform) or ""

    return row


def _write_excel_sheet(ws: Any, cols: list, rows: list[dict], *, header_font: Any, header_fill: Any) -> None:
    """Write a header row + data rows to an openpyxl worksheet."""
    from openpyxl.utils import get_column_letter
    for col_idx, key in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=key)
        cell.font = header_font
        cell.fill = header_fill
    for row_idx, record in enumerate(rows, start=2):
        for col_idx, key in enumerate(cols, start=1):
            val = record.get(key, "")
            ws.cell(row=row_idx, column=col_idx, value=val if val != "" else None)
    for col_idx, key in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(50, len(key) + 2))
    ws.freeze_panes = "A2"


def to_excel(data: dict, output_path: Path) -> None:
    """Write metadata dict to an Excel workbook in native datacoolie format.

    Produces three sheets (connections, dataflows, schema_hints) with the same
    column layout as local_use_cases.xlsx — compatible with file_provider.py.

    Raises:
        ImportError: if openpyxl is not installed.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        raise ImportError(
            "Excel export requires openpyxl. Install it with: pip install openpyxl"
        ) from None

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    kw = {"header_font": header_font, "header_fill": header_fill}

    connections = data.get("connections", [])
    if isinstance(connections, list) and connections:
        _write_excel_sheet(
            wb.create_sheet(title="connections"),
            _CONNECTION_COLS,
            [_connection_to_row(c) for c in connections if isinstance(c, dict)],
            **kw,
        )

    dataflows = data.get("dataflows", [])
    if isinstance(dataflows, list) and dataflows:
        _write_excel_sheet(
            wb.create_sheet(title="dataflows"),
            _DATAFLOW_COLS,
            [_dataflow_to_row(df) for df in dataflows if isinstance(df, dict)],
            **kw,
        )

    schema_hints = data.get("schema_hints", [])
    if isinstance(schema_hints, list) and schema_hints:
        hints_rows: list[dict] = []
        for group in schema_hints:
            if not isinstance(group, dict):
                continue
            conn_name = group.get("connection_name", "")
            table_name = group.get("table_name", "")
            schema_name = group.get("schema_name", "")
            for hint in group.get("hints", []):
                if not isinstance(hint, dict):
                    continue
                hints_rows.append({
                    "connection_name": conn_name,
                    "table_name": table_name,
                    "schema_name": schema_name,
                    "column_name": hint.get("column_name", ""),
                    "data_type": hint.get("data_type", ""),
                    "format": hint.get("format", ""),
                    "precision": hint.get("precision", ""),
                    "scale": hint.get("scale", ""),
                    "is_active": "" if hint.get("is_active") is None else hint["is_active"],
                })
        if hints_rows:
            _write_excel_sheet(wb.create_sheet(title="schema_hints"), _SCHEMA_HINTS_COLS, hints_rows, **kw)

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Convert DataCoolie metadata between JSON, YAML, and Excel."
    )
    parser.add_argument("input_file", type=Path, help="Input metadata file (JSON, YAML, or Excel).")
    parser.add_argument("--to", choices=["json", "yaml", "excel"], required=True, dest="target_format", help="Target format.")
    parser.add_argument("--output", type=Path, default=None, help="Output file path (default: same name with new extension).")
    args = parser.parse_args()

    if not args.input_file.exists():
        print(f"ERROR: File not found: {args.input_file}", file=sys.stderr)
        sys.exit(2)

    # Detect Excel input
    is_excel_input = args.input_file.suffix.lower() in (".xlsx", ".xls")

    if is_excel_input and args.target_format == "excel":
        print("ERROR: Input and output are both Excel. Use --to json or --to yaml.", file=sys.stderr)
        sys.exit(2)

    # Load input — xlsx now routed through _loaders.load_file() (native format)
    try:
        data = load_file(args.input_file)
    except (json.JSONDecodeError, yaml.YAMLError) as e:
        print(f"ERROR: Failed to parse {args.input_file}: {e}", file=sys.stderr)
        sys.exit(2)
    except (ImportError, ValueError) as e:
        print(f"ERROR: {e}", file=sys.stderr)
        sys.exit(2)

    # Convert
    if args.target_format == "json":
        output_content = to_json(data)
        default_ext = ".json"
    elif args.target_format == "yaml":
        output_content = to_yaml(data)
        default_ext = ".yaml"
    else:  # excel
        default_ext = ".xlsx"
        output_path = args.output or args.input_file.with_suffix(default_ext)
        if output_path.resolve() == args.input_file.resolve():
            print(f"ERROR: Output path is same as input. Use --output to specify a different path.", file=sys.stderr)
            sys.exit(2)
        output_path.parent.mkdir(parents=True, exist_ok=True)
        try:
            to_excel(data, output_path)
        except ImportError as e:
            print(f"ERROR: {e}", file=sys.stderr)
            sys.exit(2)
        print(f"\u2713 Converted {args.input_file} \u2192 {output_path} (excel)")
        sys.exit(0)

    # JSON / YAML path
    if args.output:
        output_path = args.output
    else:
        output_path = args.input_file.with_suffix(default_ext)

    # Avoid overwriting input
    if output_path.resolve() == args.input_file.resolve():
        print(f"ERROR: Output path is same as input. Use --output to specify a different path.", file=sys.stderr)
        sys.exit(2)

    # Write output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(output_content)

    print(f"✓ Converted {args.input_file} → {output_path} ({args.target_format})")
    sys.exit(0)


if __name__ == "__main__":
    main()
