"""Schema extraction from data files — shared parsers for local, cloud, and OneLake paths.

Supports: CSV, TSV, JSON, JSONL, Parquet, Excel, Avro.
Each parser returns a list[dict] with keys: name, data_type, ordinal (or error).
"""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path


# ---------------------------------------------------------------------------
# Type inference
# ---------------------------------------------------------------------------

def infer_json_type(value) -> str:
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "decimal"
    if isinstance(value, list):
        return "array"
    if isinstance(value, dict):
        return "object"
    return "string"


# ---------------------------------------------------------------------------
# Bytes-based parsers (OneLake DFS SDK downloads, in-memory content)
# ---------------------------------------------------------------------------

def schema_from_bytes(content: bytes, ext: str) -> list[dict]:
    """Extract schema from raw file bytes."""
    try:
        if ext in (".csv", ".tsv"):
            text = content.decode("utf-8", errors="replace")
            reader = csv.reader(io.StringIO(text), delimiter="\t" if ext == ".tsv" else ",")
            headers = next(reader, [])
            return [{"name": h.strip(), "data_type": "string", "ordinal": i + 1}
                    for i, h in enumerate(headers) if h.strip()]
        if ext == ".json":
            data = json.loads(content)
            record = data[0] if isinstance(data, list) and data else data if isinstance(data, dict) else {}
            return [{"name": k, "data_type": infer_json_type(v), "ordinal": i + 1}
                    for i, (k, v) in enumerate(record.items())]
        if ext == ".jsonl":
            first_line = content.split(b"\n", 1)[0].decode("utf-8", errors="replace")
            record = json.loads(first_line) if first_line.strip() else {}
            return [{"name": k, "data_type": infer_json_type(v), "ordinal": i + 1}
                    for i, (k, v) in enumerate(record.items())]
        if ext == ".parquet":
            return schema_from_parquet_fileobj(io.BytesIO(content))
        if ext in (".xlsx", ".xls"):
            return schema_from_excel_fileobj(io.BytesIO(content))
        if ext == ".avro":
            return schema_from_avro_fileobj(io.BytesIO(content))
    except Exception as e:
        return [{"error": str(e)}]
    return []


# ---------------------------------------------------------------------------
# File-object / path-based parsers
# ---------------------------------------------------------------------------

def schema_from_csv_path(file_path: Path, delimiter: str = ",") -> list[dict]:
    try:
        with open(file_path, encoding="utf-8", newline="") as f:
            reader = csv.reader(f, delimiter=delimiter)
            headers = next(reader, [])
        return [{"name": h.strip(), "data_type": "string", "ordinal": i + 1}
                for i, h in enumerate(headers) if h.strip()]
    except Exception as e:
        return [{"error": str(e)}]


def schema_from_json_path(file_path: Path) -> list[dict]:
    try:
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)
        record = data[0] if isinstance(data, list) and data else data if isinstance(data, dict) else {}
        return [{"name": k, "data_type": infer_json_type(v), "ordinal": i + 1}
                for i, (k, v) in enumerate(record.items())]
    except Exception as e:
        return [{"error": str(e)}]


def schema_from_jsonl_path(file_path: Path) -> list[dict]:
    try:
        with open(file_path, encoding="utf-8") as f:
            first_line = f.readline()
        record = json.loads(first_line) if first_line.strip() else {}
        return [{"name": k, "data_type": infer_json_type(v), "ordinal": i + 1}
                for i, (k, v) in enumerate(record.items())]
    except Exception as e:
        return [{"error": str(e)}]


def schema_from_parquet_fileobj(fileobj) -> list[dict]:
    """Read Parquet schema from a path or file-like object (reads footer only)."""
    try:
        import pyarrow.parquet as pq
        schema = pq.read_schema(fileobj)
        return [{"name": field.name, "data_type": str(field.type), "ordinal": i + 1}
                for i, field in enumerate(schema)]
    except ImportError:
        return [{"error": "pyarrow not installed — pip install pyarrow"}]
    except Exception as e:
        return [{"error": str(e)}]


def schema_from_excel_path(file_path: Path) -> list[dict]:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        wb.close()
        return [{"name": str(h), "data_type": "string", "ordinal": i + 1}
                for i, h in enumerate(headers) if h is not None]
    except ImportError:
        return [{"error": "openpyxl not installed — pip install openpyxl"}]
    except Exception as e:
        return [{"error": str(e)}]


def schema_from_excel_fileobj(fileobj) -> list[dict]:
    try:
        import openpyxl
        data = fileobj.read() if hasattr(fileobj, "read") else fileobj
        wb = openpyxl.load_workbook(io.BytesIO(data) if isinstance(data, bytes) else fileobj, read_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
        wb.close()
        return [{"name": str(h), "data_type": "string", "ordinal": i + 1}
                for i, h in enumerate(headers) if h is not None]
    except ImportError:
        return [{"error": "openpyxl not installed — pip install openpyxl"}]
    except Exception as e:
        return [{"error": str(e)}]


def schema_from_avro_fileobj(fileobj) -> list[dict]:
    try:
        import fastavro
        reader = fastavro.reader(fileobj)
        writer_schema = reader.writer_schema
        if isinstance(writer_schema, dict) and "fields" in writer_schema:
            return [{"name": field["name"], "data_type": str(field.get("type", "string")), "ordinal": i + 1}
                    for i, field in enumerate(writer_schema["fields"])]
        return []
    except ImportError:
        return [{"error": "fastavro not installed — pip install fastavro"}]
    except Exception as e:
        return [{"error": str(e)}]


# ---------------------------------------------------------------------------
# fsspec cloud dispatcher
# ---------------------------------------------------------------------------

def schema_from_cloud_file(fs, path: str, ext: str) -> list[dict]:
    """Extract schema from a cloud file via fsspec filesystem."""
    try:
        if ext in (".csv", ".tsv"):
            with fs.open(path, "r", encoding="utf-8", newline="") as f:
                reader = csv.reader(f, delimiter="\t" if ext == ".tsv" else ",")
                headers = next(reader, [])
            return [{"name": h.strip(), "data_type": "string", "ordinal": i + 1}
                    for i, h in enumerate(headers) if h.strip()]
        if ext == ".json":
            with fs.open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            record = data[0] if isinstance(data, list) and data else data if isinstance(data, dict) else {}
            return [{"name": k, "data_type": infer_json_type(v), "ordinal": i + 1}
                    for i, (k, v) in enumerate(record.items())]
        if ext == ".jsonl":
            with fs.open(path, "r", encoding="utf-8") as f:
                first_line = f.readline()
            record = json.loads(first_line) if first_line.strip() else {}
            return [{"name": k, "data_type": infer_json_type(v), "ordinal": i + 1}
                    for i, (k, v) in enumerate(record.items())]
        if ext == ".parquet":
            with fs.open(path, "rb") as f:
                return schema_from_parquet_fileobj(f)
        if ext in (".xlsx", ".xls"):
            with fs.open(path, "rb") as f:
                return schema_from_excel_fileobj(f)
        if ext == ".avro":
            with fs.open(path, "rb") as f:
                return schema_from_avro_fileobj(f)
    except Exception as e:
        return [{"error": str(e)}]
    return []


# ---------------------------------------------------------------------------
# Local path dispatcher
# ---------------------------------------------------------------------------

def schema_from_local_file(file_path: Path) -> list[dict]:
    """Extract schema from a local file by extension."""
    ext = file_path.suffix.lower()
    if ext in (".csv", ".tsv"):
        return schema_from_csv_path(file_path, "\t" if ext == ".tsv" else ",")
    if ext == ".json":
        return schema_from_json_path(file_path)
    if ext == ".jsonl":
        return schema_from_jsonl_path(file_path)
    if ext == ".parquet":
        return schema_from_parquet_fileobj(file_path)
    if ext in (".xlsx", ".xls"):
        return schema_from_excel_path(file_path)
    if ext == ".avro":
        with open(file_path, "rb") as f:
            return schema_from_avro_fileobj(f)
    return []
