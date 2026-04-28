"""File-based metadata provider — reads connections, dataflows, and schema hints
from YAML, JSON, or Excel (.xlsx / .xls) configuration files.

``FileProvider`` is the primary standalone / development metadata backend.
Watermark state is stored as JSON files on the platform's file system.

Supported formats
-----------------

**YAML / JSON** — hierarchical, nested structure:

.. code-block:: yaml

    connections:
      - name: bronze_adls
        connection_type: file
        format: delta
        configure:
          base_path: abfss://bronze@storage/
          use_schema_hint: true

      - name: source_erp
        connection_type: database
        format: jdbc
        database: ERP
        configure:
          host: erp-server.example.com
          port: 1433
          username: etl_reader

    dataflows:
      - name: orders_bronze_to_silver
        stage: bronze2silver
        source:
          connection_name: bronze_adls
          table: orders
          watermark_columns: [modified_at]
        destination:
          connection_name: silver_lakehouse
          table: dim_orders
          load_type: merge_upsert
          merge_keys: [order_id]
        transform:
          schema_hints:
            - column_name: amount
              data_type: DECIMAL
              precision: 18
              scale: 2

    schema_hints:
      - connection_name: bronze_adls
        table_name: orders
        hints:
          - column_name: order_date
            data_type: DATE
            format: yyyy-MM-dd

**Excel (.xlsx / .xls)** — flat workbook with three sheets:

*connections* sheet — one row per connection:

  Required: ``name``, ``connection_type``
  Optional: ``connection_id``, ``format``, ``catalog``, ``database``, ``is_active``
  Nested ``configure`` via ``configure_*`` columns (e.g. ``configure_base_path``, ``configure_host``,
    ``configure_use_schema_hint``)
  ``catalog`` and ``database`` are top-level columns (not ``configure_catalog`` or ``configure_database``)
  ``secrets_ref``: JSON column mapping config field names to vault key names
    (e.g. ``{"password": "vault/db-pass", "api_key": "vault/api-key"}``)

*dataflows* sheet — one row per dataflow:

  Required: ``source_connection_name``, ``source_table``,
    ``destination_connection_name``, ``destination_table``
  Optional: ``dataflow_id``, ``name``, ``stage``, ``description``, ``group_number``,
    ``execution_order``, ``processing_mode``, ``is_active``
  ``configure``: JSON column for the dataflow's own configure dict.
  ``transform``: JSON column for the full transform dict (individual ``transform_*``
    columns take precedence when both are present).
  Source fields prefixed ``source_``; destination fields prefixed ``destination_``;
  List columns (e.g. ``source_watermark_columns``, ``destination_merge_keys``,
  ``destination_partition_columns``) accept comma-separated values.
  ``transform_schema_hints``, ``transform_additional_columns``,
  ``transform_partition_columns`` accept JSON strings.

*schema_hints* sheet — one row per hint (grouped internally by
  ``connection_name`` + ``table_name`` + optional ``schema_name``):

  Columns: ``connection_name``, ``table_name``, ``schema_name`` (optional),
    ``column_name``, ``data_type``, ``precision``, ``scale``, ``format``

**Separate files** — each section may live in its own file instead of (or
in addition to) the primary ``config_path``.  Pass any combination of:

* ``connections_path`` — file that contains only a ``connections`` list.
* ``schema_hints_path`` — file that contains only a ``schema_hints`` list.

Each override file can be YAML, JSON, or Excel in the same format as the
corresponding section in the primary file.  When a separate path is
specified, its section **replaces** the same section from the primary file.
"""

from __future__ import annotations

import json
from typing import Any, Dict, List, Optional, Tuple

from datacoolie.core.constants import WATERMARK_FILE_NAME
from datacoolie.core.exceptions import MetadataError, WatermarkError
from datacoolie.core.models import (
    Connection,
    DataFlow,
    Destination,
    SchemaHint,
    Source,
    Transform,
)
from datacoolie.metadata.base import BaseMetadataProvider
from datacoolie.platforms.base import BasePlatform
from datacoolie.utils.converters import convert_to_bool, convert_to_int, parse_json
from datacoolie.utils.helpers import ensure_list
from datacoolie.utils.path_utils import normalize_path


class FileProvider(BaseMetadataProvider):
    """Metadata provider backed by YAML, JSON, or Excel configuration files.

    Args:
        config_path: Path to the primary YAML, JSON, or Excel metadata file.
            May contain ``connections``, ``dataflows``, and/or
            ``schema_hints`` sections.
        platform: Platform used for watermark file I/O.
        connections_path: Optional separate file that provides the
            ``connections`` list.  Overrides any ``connections`` section
            in *config_path* when supplied.
        schema_hints_path: Optional separate file that provides the
            ``schema_hints`` list.  Overrides any ``schema_hints`` section
            in *config_path* when supplied.
        watermark_base_path: Root directory for watermark files.
            Defaults to ``<config_path_parent>/watermarks``.
        enable_cache: Enable the in-memory cache.
    """

    _EXCEL_LIST_COLS: frozenset = frozenset({
        "source_watermark_columns",
        "destination_merge_keys",
        "destination_partition_columns",
    })
    _EXCEL_JSON_COLS: frozenset = frozenset({
        "configure",
        "source_configure",
        "destination_configure",
        "transform_deduplicate_columns",
        "transform_latest_data_columns",
        "transform_additional_columns",
        "transform_schema_hints",
        "transform_configure",
    })

    def __init__(
        self,
        config_path: str,
        platform: BasePlatform,
        *,
        connections_path: Optional[str] = None,
        schema_hints_path: Optional[str] = None,
        watermark_base_path: Optional[str] = None,
        enable_cache: bool = True,
        eager_prefetch: bool = False,
    ) -> None:
        super().__init__(enable_cache=enable_cache, eager_prefetch=eager_prefetch)
        self._config_path = normalize_path(config_path)
        self._connections_path = normalize_path(connections_path) if connections_path else None
        self._schema_hints_path = normalize_path(schema_hints_path) if schema_hints_path else None
        self._platform = platform
        config_parent = self._config_path.rsplit("/", 1)[0] if "/" in self._config_path else ""
        default_watermark_base = f"{config_parent}/watermarks" if config_parent else "watermarks"
        self._watermark_base_path = normalize_path(watermark_base_path or default_watermark_base)
        self._data: Dict[str, Any] = {}
        self._load_config()
        # Memoised outputs of the ``_build_*`` helpers — the underlying
        # ``self._data`` dict is immutable during the provider's
        # lifetime, so the parsed models can be cached.
        self._all_connections_cache: Optional[List[Connection]] = None
        self._active_connections_cache: Optional[List[Connection]] = None
        # name -> Connection lookup for O(1) ``_resolve_connection``.
        self._connection_by_name: Optional[Dict[str, Connection]] = None
        # Full (active + inactive) dataflow list; filtered variants are
        # derived on demand from this list.
        self._all_dataflows_cache: Optional[List[DataFlow]] = None
        self._maybe_eager_prefetch()

    # ------------------------------------------------------------------
    # Config loading
    # ------------------------------------------------------------------

    def _load_file(self, path: str) -> Dict[str, Any]:
        """Load and parse a single YAML, JSON, or Excel file.

        Excel files are read directly from the filesystem (openpyxl requires
        a seekable binary stream); all other formats go through the platform.
        """
        path_lower = path.lower()
        if path_lower.endswith((".xlsx", ".xls")):
            try:
                return self._parse_excel(path)
            except MetadataError:
                raise
            except Exception as exc:
                raise MetadataError(f"Cannot parse metadata config: {path}") from exc

        try:
            raw = self._platform.read_file(path)
        except Exception as exc:
            raise MetadataError(f"Cannot read metadata config: {path}") from exc

        try:
            if path_lower.endswith((".yaml", ".yml")):
                return self._parse_yaml(raw)
            return json.loads(raw)
        except MetadataError:
            raise
        except Exception as exc:
            raise MetadataError(f"Cannot parse metadata config: {path}") from exc

    def _load_config(self) -> None:
        """Load the primary config file then overlay any separate section files."""
        self._data = self._load_file(self._config_path)

        # Separate files override the corresponding section from the primary file.
        if self._connections_path:
            overlay = self._load_file(self._connections_path)
            self._data["connections"] = overlay.get("connections", [])

        if self._schema_hints_path:
            overlay = self._load_file(self._schema_hints_path)
            self._data["schema_hints"] = overlay.get("schema_hints", [])

    @staticmethod
    def _parse_yaml(raw: str) -> Dict[str, Any]:
        """Parse YAML text, importing PyYAML lazily."""
        try:
            import yaml  # noqa: WPS433 — optional dependency
        except ImportError as exc:
            raise MetadataError(
                "PyYAML is required for YAML metadata files.  "
                "Install it with:  pip install pyyaml"
            ) from exc
        result = yaml.safe_load(raw)
        if result is None:
            return {}
        if not isinstance(result, dict):
            raise MetadataError("YAML metadata file must contain a mapping at root level")
        return result


    # -- Excel cell helpers -------------------------------------------------

    @staticmethod
    def _excel_sheet_rows(wb: Any, sheet_name: str) -> List[Dict[str, Any]]:
        """Return a list of row-dicts for *sheet_name*, or ``[]`` if absent."""
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        rows: List[Dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row):
                continue
            rows.append({headers[i]: row[i] for i in range(len(headers))})
        return rows

    @staticmethod
    def _cast(value: Any) -> Any:
        """Normalise cell values: strip strings, return ``None`` for empty."""
        if isinstance(value, str):
            value = value.strip()
            return value if value != "" else None
        return value

    @staticmethod
    def _safe_bool(value: Any) -> bool:
        """Coerce a cell to ``bool``; returns ``False`` for unrecognised values."""
        try:
            return convert_to_bool(value)
        except ValueError:
            return False

    @staticmethod
    def _json_cell(value: Any) -> Any:
        """Parse a JSON object/array cell; return ``None`` if blank."""
        if value is None:
            return None
        s = str(value).strip() if isinstance(value, str) else value
        try:
            result = parse_json(s, raise_on_error=True)
        except ValueError as exc:
            raise MetadataError(f"Invalid JSON cell value: {value!r}") from exc
        return result if result else None

    # -- Excel parsing ------------------------------------------------------

    @classmethod
    def _parse_excel(cls, path: str) -> Dict[str, Any]:
        """Parse an Excel workbook (.xlsx/.xls) into the standard config dict.

        Expected sheets: ``connections``, ``dataflows``, ``schema_hints``
        (all optional — missing sheets are treated as empty lists).

        See module docstring for the per-sheet column conventions.
        """
        try:
            import openpyxl  # noqa: WPS433 — optional dependency
        except ImportError as exc:
            raise MetadataError(
                "openpyxl is required for Excel metadata files.  "
                "Install it with:  pip install openpyxl"
            ) from exc
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise MetadataError(f"Cannot read Excel metadata file: {path}") from exc
        try:
            return {
                "connections": cls._parse_excel_connections(wb),
                "dataflows": cls._parse_excel_dataflows(wb),
                "schema_hints": cls._parse_excel_schema_hints(wb),
            }
        finally:
            wb.close()

    @classmethod
    def _parse_excel_connections(cls, wb: Any) -> List[Dict[str, Any]]:
        connections: List[Dict[str, Any]] = []
        for row in cls._excel_sheet_rows(wb, "connections"):
            conn: Dict[str, Any] = {}
            cfg: Dict[str, Any] = {}
            for key, raw_val in row.items():
                val = cls._cast(raw_val)
                if key == "configure":
                    parsed = cls._json_cell(raw_val)
                    if isinstance(parsed, dict):
                        cfg.update(parsed)
                elif key.startswith("configure_"):
                    sub = key[len("configure_"):]
                    if val is not None:
                        cfg[sub] = val
                elif key == "secrets_ref":
                    parsed = cls._json_cell(raw_val)
                    if parsed is not None:
                        conn[key] = parsed
                elif key == "is_active":
                    if val is not None:
                        conn[key] = cls._safe_bool(val)
                elif val is not None:
                    conn[key] = val
            if cfg:
                conn["configure"] = cfg
            if conn:
                connections.append(conn)
        return connections

    @classmethod
    def _parse_excel_dataflows(cls, wb: Any) -> List[Dict[str, Any]]:
        dataflows: List[Dict[str, Any]] = []
        for row in cls._excel_sheet_rows(wb, "dataflows"):
            df_dict: Dict[str, Any] = {}
            src: Dict[str, Any] = {}
            dest: Dict[str, Any] = {}
            transform: Dict[str, Any] = {}
            transform_base: Dict[str, Any] = {}
            for key, raw_val in row.items():
                if key == "transform":
                    parsed = cls._json_cell(raw_val)
                    if isinstance(parsed, dict):
                        transform_base = parsed
                    continue
                if key == "is_active":
                    parsed_is_active = cls._cast(raw_val)
                    if parsed_is_active is not None:
                        df_dict["is_active"] = cls._safe_bool(parsed_is_active)
                    continue
                if key.startswith("source_"):
                    target, sub = src, key[len("source_"):]
                elif key.startswith("destination_"):
                    target, sub = dest, key[len("destination_"):]
                elif key.startswith("transform_"):
                    target, sub = transform, key[len("transform_"):]
                else:
                    val = cls._json_cell(raw_val) if key in cls._EXCEL_JSON_COLS else cls._cast(raw_val)
                    if val is not None:
                        df_dict[key] = val
                    continue
                if key in cls._EXCEL_LIST_COLS:
                    parsed = ensure_list(raw_val)
                    if parsed:
                        target[sub] = parsed
                elif key in cls._EXCEL_JSON_COLS:
                    parsed_json = cls._json_cell(raw_val)
                    if parsed_json is not None:
                        target[sub] = parsed_json
                else:
                    val = cls._cast(raw_val)
                    if val is not None:
                        target[sub] = val
            if src:
                df_dict["source"] = src
            if dest:
                df_dict["destination"] = dest
            merged_transform = {**transform_base, **transform}  # transform_* sub-keys win
            if merged_transform:
                df_dict["transform"] = merged_transform
            if df_dict:
                dataflows.append(df_dict)
        return dataflows

    @classmethod
    def _parse_excel_schema_hints(cls, wb: Any) -> List[Dict[str, Any]]:
        hints_map: Dict[tuple, Dict[str, Any]] = {}
        for row in cls._excel_sheet_rows(wb, "schema_hints"):
            conn_name = cls._cast(row.get("connection_name"))
            table_name = cls._cast(row.get("table_name"))
            schema_name = cls._cast(row.get("schema_name"))
            if not conn_name or not table_name:
                continue
            group_key = (conn_name, table_name, schema_name)
            if group_key not in hints_map:
                hints_map[group_key] = {
                    "connection_name": conn_name,
                    "table_name": table_name,
                    "hints": [],
                }
                if schema_name is not None:
                    hints_map[group_key]["schema_name"] = schema_name
            hint: Dict[str, Any] = {}
            for field in ("column_name", "data_type", "format"):
                val = cls._cast(row.get(field))
                if val is not None:
                    hint[field] = val
            for field in ("precision", "scale"):
                converted = convert_to_int(row.get(field))
                if converted is not None:
                    hint[field] = converted
            default_value = cls._cast(row.get("default_value"))
            if default_value is not None:
                hint["default_value"] = default_value
            ordinal_position = convert_to_int(row.get("ordinal_position"))
            if ordinal_position is not None:
                hint["ordinal_position"] = ordinal_position
            parsed_is_active = cls._cast(row.get("is_active"))
            if parsed_is_active is not None:
                hint["is_active"] = cls._safe_bool(parsed_is_active)
            if hint:
                hints_map[group_key]["hints"].append(hint)
        return list(hints_map.values())

    # ------------------------------------------------------------------
    # Connection helpers
    # ------------------------------------------------------------------

    def _build_connections(self, *, active_only: bool = True) -> List[Connection]:
        """Build ``Connection`` models from the ``connections`` section.

        Results are memoised — the source dict is immutable after
        ``_load_config`` — so repeated calls are O(1).
        """
        if self._all_connections_cache is None:
            raw_list: List[Dict[str, Any]] = self._data.get("connections", [])
            built: List[Connection] = []
            for raw in raw_list:
                try:
                    built.append(Connection(**raw))
                except Exception as exc:
                    raise MetadataError(
                        f"Invalid connection definition: {raw.get('name', '?')}"
                    ) from exc
            self._all_connections_cache = built
            self._active_connections_cache = [c for c in built if c.is_active]
            self._connection_by_name = {c.name: c for c in built}
        return (
            self._active_connections_cache if active_only  # type: ignore[return-value]
            else self._all_connections_cache
        )

    def _resolve_connection(self, name_or_ref: str | Dict[str, Any]) -> Connection:
        """Resolve a connection name or inline dict to a ``Connection``.

        Uses the memoised ``name -> Connection`` map for O(1) lookup.
        """
        if isinstance(name_or_ref, dict):
            return Connection(**name_or_ref)
        # Warm the memoised map if not yet built.
        if self._connection_by_name is None:
            self._build_connections(active_only=False)
        conn = self._connection_by_name.get(name_or_ref) if self._connection_by_name else None
        if conn is None:
            raise MetadataError(f"Connection not found: {name_or_ref}")
        return conn

    # ------------------------------------------------------------------
    # Dataflow helpers
    # ------------------------------------------------------------------

    def _build_dataflows(
        self,
        *,
        stages: Optional[List[str]] = None,
        active_only: bool = True,
    ) -> List[DataFlow]:
        """Build ``DataFlow`` models from the ``dataflows`` section.

        The full (active + inactive, unfiltered) list is built once and
        memoised; filtered calls derive from it in memory.
        """
        if self._all_dataflows_cache is None:
            raw_list: List[Dict[str, Any]] = self._data.get("dataflows", [])
            built: List[DataFlow] = []
            for raw in raw_list:
                try:
                    built.append(self._build_single_dataflow(raw))
                except MetadataError:
                    raise
                except Exception as exc:
                    raise MetadataError(
                        f"Invalid dataflow definition: {raw.get('name', '?')}"
                    ) from exc
            self._all_dataflows_cache = built

        dataflows = self._all_dataflows_cache
        if active_only:
            dataflows = [df for df in dataflows if df.is_active]
        if stages is not None:
            stage_set = set(stages)
            dataflows = [df for df in dataflows if df.stage in stage_set]
        return dataflows

    def _build_single_dataflow(self, raw: Dict[str, Any]) -> DataFlow:
        """Build one ``DataFlow``, resolving connection references."""
        # -- source --
        src_raw: Dict[str, Any] = dict(raw.get("source", {}))
        src_conn_ref = src_raw.pop("connection_name", src_raw.pop("connection", None))
        if src_conn_ref is None:
            raise MetadataError(
                f"Dataflow '{raw.get('name', '?')}' source must have "
                "'connection_name' or 'connection'"
            )
        source = Source(connection=self._resolve_connection(src_conn_ref), **src_raw)

        # -- destination --
        dest_raw: Dict[str, Any] = dict(raw.get("destination", {}))
        dest_conn_ref = dest_raw.pop("connection_name", dest_raw.pop("connection", None))
        if dest_conn_ref is None:
            raise MetadataError(
                f"Dataflow '{raw.get('name', '?')}' destination must have "
                "'connection_name' or 'connection'"
            )
        destination = Destination(connection=self._resolve_connection(dest_conn_ref), **dest_raw)

        # -- transform (optional) --
        transform_raw: Dict[str, Any] = raw.get("transform", {})

        return DataFlow(
            source=source,
            destination=destination,
            transform=Transform(**transform_raw) if transform_raw else Transform(),
            **{k: v for k, v in raw.items() if k not in ("source", "destination", "transform")},
        )

    # ------------------------------------------------------------------
    # Schema hint helpers
    # ------------------------------------------------------------------

    def _build_schema_hints(
        self,
        connection_id: str,
        table_name: str,
        schema_name: Optional[str] = None,
    ) -> List[SchemaHint]:
        """Build ``SchemaHint`` models matching *connection_id* and *table_name*."""
        raw_list: List[Dict[str, Any]] = self._data.get("schema_hints", [])
        # All names/ids that identify the target connection (for flexible matching).
        all_conns = self._build_connections(active_only=False)
        target_refs = {connection_id} | {c.name for c in all_conns if c.connection_id == connection_id}

        results: List[SchemaHint] = []
        for group in raw_list:
            conn_ref = group.get("connection_name") or group.get("connection_id")
            if conn_ref not in target_refs:
                continue
            if group.get("table_name", "").lower() != table_name.lower():
                continue
            group_schema = group.get("schema_name")
            if schema_name is not None and group_schema is not None:
                if group_schema.lower() != schema_name.lower():
                    continue
            for hint_raw in group.get("hints", []):
                try:
                    results.append(SchemaHint(**hint_raw))
                except Exception as exc:
                    raise MetadataError(
                        f"Invalid schema hint for {table_name}: {hint_raw}"
                    ) from exc
        return results

    # ------------------------------------------------------------------
    # Watermark file I/O
    # ------------------------------------------------------------------

    def _watermark_path(self, dataflow_id: str) -> str:
        """Build the watermark file path for a dataflow.

        The folder segment is ``{stage}_{name}_{dataflow_id}`` when both
        *stage* and *name* are set on the dataflow, otherwise plain
        *dataflow_id* is used (backward-compatible fallback).

        Uses :meth:`get_dataflow_by_id` so the base-class cache is
        consulted first — avoids rebuilding the full dataflow list on
        every watermark read/write after a bulk-load.
        """
        df = self.get_dataflow_by_id(dataflow_id, attach_schema_hints=False)
        if df is not None:
            parts = []
            if df.stage:
                parts.append(df.stage)
            if df.name:
                parts.append(df.name)
            parts.append(dataflow_id)
            folder = "_".join(parts)
        else:
            folder = dataflow_id
        return f"{self._watermark_base_path.rstrip('/')}/{folder}/{WATERMARK_FILE_NAME}"

    # ------------------------------------------------------------------
    # Abstract method implementations
    # ------------------------------------------------------------------

    def _bulk_load(
        self,
    ) -> Tuple[
        List[Connection],
        List[DataFlow],
        Dict[Tuple[str, Optional[str], str], List[SchemaHint]],
    ]:
        """Bulk-load all metadata from the in-memory ``self._data``.

        Walks ``self._data['schema_hints']`` once and groups by
        ``(connection_id, schema_name, table_name)`` so that the
        per-dataflow attach loop becomes a pure cache hit.
        """
        connections = self._build_connections(active_only=False)
        dataflows = self._build_dataflows(active_only=False)

        # Build name -> connection_id map for resolving hint groups that
        # reference ``connection_name`` rather than ``connection_id``.
        name_to_id: Dict[str, str] = {c.name: c.connection_id for c in connections}
        valid_ids: set = {c.connection_id for c in connections}

        grouped: Dict[Tuple[str, Optional[str], str], List[SchemaHint]] = {}
        for group in self._data.get("schema_hints", []):
            conn_ref = group.get("connection_name") or group.get("connection_id")
            table = group.get("table_name")
            if not conn_ref or not table:
                continue
            if conn_ref in valid_ids:
                conn_id = conn_ref
            else:
                conn_id = name_to_id.get(conn_ref)
            if not conn_id:
                continue
            schema = group.get("schema_name") or None  # normalise '' -> None
            key = (conn_id, schema, table)
            for hint_raw in group.get("hints", []):
                try:
                    grouped.setdefault(key, []).append(SchemaHint(**hint_raw))
                except Exception as exc:
                    raise MetadataError(
                        f"Invalid schema hint for {table}: {hint_raw}"
                    ) from exc
        return connections, dataflows, grouped

    def _fetch_connections(self, *, active_only: bool = True) -> List[Connection]:
        return self._build_connections(active_only=active_only)

    def _fetch_connection_by_id(self, connection_id: str) -> Optional[Connection]:
        return next(
            (c for c in self._build_connections(active_only=False) if c.connection_id == connection_id),
            None,
        )

    def _fetch_connection_by_name(self, name: str) -> Optional[Connection]:
        return next(
            (c for c in self._build_connections(active_only=False) if c.name == name),
            None,
        )

    def _fetch_dataflows(
        self,
        *,
        stages: Optional[List[str]] = None,
        active_only: bool = True,
    ) -> List[DataFlow]:
        return self._build_dataflows(stages=stages, active_only=active_only)

    def _fetch_dataflow_by_id(self, dataflow_id: str) -> Optional[DataFlow]:
        return next(
            (df for df in self._build_dataflows(active_only=False) if df.dataflow_id == dataflow_id),
            None,
        )

    def get_watermark(self, dataflow_id: str) -> Optional[str]:
        """Return the raw watermark JSON string for *dataflow_id*, or ``None``."""
        path = self._watermark_path(dataflow_id)
        try:
            if not self._platform.file_exists(path):
                return None
            raw = self._platform.read_file(path)
            return raw if raw and raw.strip() else None
        except Exception as exc:
            raise WatermarkError(f"Cannot read watermark at {path}") from exc

    def update_watermark(
        self,
        dataflow_id: str,
        watermark_value: str,
        *,
        job_id: Optional[str] = None,
        dataflow_run_id: Optional[str] = None,
    ) -> None:
        """Write the serialised watermark JSON to a file."""
        path = self._watermark_path(dataflow_id)
        try:
            self._platform.write_file(path, watermark_value, overwrite=True)
        except Exception as exc:
            raise WatermarkError(f"Cannot write watermark at {path}") from exc

    def _fetch_schema_hints(
        self,
        connection_id: str,
        table_name: str,
        schema_name: Optional[str] = None,
    ) -> List[SchemaHint]:
        return self._build_schema_hints(
            connection_id=connection_id,
            table_name=table_name,
            schema_name=schema_name,
        )

    def close(self) -> None:
        """Clear cache, memoised builders, and internal data."""
        super().close()
        self._data.clear()
        self._all_connections_cache = None
        self._active_connections_cache = None
        self._connection_by_name = None
        self._all_dataflows_cache = None
