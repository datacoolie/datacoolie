"""Tests for datacoolie.engines.polars_engine — requires polars + deltalake."""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict
from unittest.mock import MagicMock, patch

import pytest

pl = pytest.importorskip("polars", reason="polars not installed")
deltalake = pytest.importorskip("deltalake", reason="deltalake not installed")

from datacoolie.core.constants import DEFAULT_AUTHOR, FileInfoColumn, SystemColumn  # noqa: E402
from datacoolie.core.exceptions import EngineError  # noqa: E402
from datacoolie.engines.polars_engine import PolarsEngine  # noqa: E402
from datacoolie.platforms.base import FileInfo  # noqa: E402


# =====================================================================
# Fixtures
# =====================================================================


@pytest.fixture()
def engine() -> PolarsEngine:
    return PolarsEngine()


@pytest.fixture()
def sample_lf() -> pl.LazyFrame:
    return pl.DataFrame(
        {
            "id": [1, 2, 3],
            "name": ["Alice", "Bob", "Carol"],
            "date_str": ["2026-01-01", "2026-01-02", "2026-01-03"],
        }
    ).lazy()


@pytest.fixture()
def delta_path(tmp_path: Path) -> Path:
    """Write a Delta table and return the path."""
    path = tmp_path / "delta_table"
    eager = pl.DataFrame(
        {
            "id": [1, 2, 3],
            "name": ["Alice", "Bob", "Carol"],
            "date_str": ["2026-01-01", "2026-01-02", "2026-01-03"],
        }
    )
    eager.write_delta(str(path))
    return path


# =====================================================================
# Read
# =====================================================================


class TestReadParquet:
    def test_read_parquet(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = tmp_path / "data.parquet"
        sample_lf.collect().write_parquet(str(path))
        result = engine.read_parquet(str(path))
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert collected.height == 3
        # __file_path is embedded by include_file_paths at scan time
        assert {"id", "name", "date_str"}.issubset(set(collected.columns))
        assert FileInfoColumn.FILE_PATH in collected.columns


class TestReadDelta:
    def test_read_delta(self, engine: PolarsEngine, delta_path: Path) -> None:
        result = engine.read_delta(str(delta_path))
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 3


class TestReadCsv:
    def test_read_csv(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = tmp_path / "data.csv"
        sample_lf.collect().write_csv(str(path))
        result = engine.read_csv(str(path))
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 3


class TestReadJson:
    def test_read_json(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = tmp_path / "data.json"
        sample_lf.collect().write_json(str(path))
        result = engine.read_json(str(path))
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 3


class TestReadExcel:
    def test_read_excel_single_file(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = tmp_path / "data.xlsx"
        sample_lf.collect().write_excel(str(path))
        result = engine.read_excel(str(path))
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 3

    def test_read_excel_injects_file_path(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = tmp_path / "data.xlsx"
        sample_lf.collect().write_excel(str(path))
        result = engine.read_excel(str(path)).collect()
        assert FileInfoColumn.FILE_PATH in result.columns

    def test_read_excel_multi_file(self, engine: PolarsEngine, tmp_path: Path) -> None:
        p1 = tmp_path / "a.xlsx"
        p2 = tmp_path / "b.xlsx"
        pl.DataFrame({"id": [1]}).write_excel(str(p1))
        pl.DataFrame({"id": [2]}).write_excel(str(p2))
        out = engine.read_excel([str(p1), str(p2)]).collect()
        assert out.height == 2
        assert FileInfoColumn.FILE_PATH in out.columns

    def test_read_excel_no_files_raises(self, engine: PolarsEngine, tmp_path: Path) -> None:
        with pytest.raises(FileNotFoundError):
            engine.read_excel(str(tmp_path / "nonexistent.xlsx"))

    def test_read_excel_directory(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """read_excel resolves .xlsx files from a directory."""
        from datacoolie.platforms.local_platform import LocalPlatform
        engine.set_platform(LocalPlatform())
        p1 = tmp_path / "a.xlsx"
        p2 = tmp_path / "b.xlsx"
        pl.DataFrame({"x": [10]}).write_excel(str(p1))
        pl.DataFrame({"x": [20]}).write_excel(str(p2))
        out = engine.read_excel(str(tmp_path)).collect()
        assert out.height == 2

    def test_read_excel_with_sheet_id(self, engine: PolarsEngine, tmp_path: Path) -> None:
        path = tmp_path / "data.xlsx"
        pl.DataFrame({"v": [1, 2]}).write_excel(str(path))
        result = engine.read_excel(str(path), options={"sheet_id": 1})
        assert result.collect().height == 2

    def test_read_excel_multi_sheet(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """Multi-sheet dict return is flattened into a single DataFrame."""
        path = tmp_path / "multi.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["id"])
        ws1.append([1])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["id"])
        ws2.append([2])
        wb.save(str(path))
        # sheet_id=0 tells Polars to read all sheets → returns dict
        result = engine.read_excel(str(path), options={"sheet_id": 0})
        collected = result.collect()
        assert collected.height == 2

    def test_read_excel_default_first_sheet_only(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """Without sheet_id/sheet_name, only the first sheet is read."""
        path = tmp_path / "multi.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1.append(["id"])
        ws1.append([1])
        ws1.append([2])
        ws2 = wb.create_sheet("Second")
        ws2.append(["id"])
        ws2.append([10])
        wb.save(str(path))
        result = engine.read_excel(str(path)).collect()
        # Default reads only first sheet (2 rows), NOT both sheets (3 rows)
        assert result.height == 2

    def test_read_excel_with_sheet_name(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """sheet_name option selects a specific sheet by name."""
        path = tmp_path / "named.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Alpha"
        ws1.append(["val"])
        ws1.append([100])
        ws2 = wb.create_sheet("Beta")
        ws2.append(["val"])
        ws2.append([200])
        ws2.append([300])
        wb.save(str(path))
        result = engine.read_excel(str(path), options={"sheet_name": "Beta"}).collect()
        assert result.height == 2
        assert result["val"].to_list() == [200, 300]

    def test_read_excel_file_path_values_correct(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """__file_path column contains the actual file paths."""
        p1 = tmp_path / "a.xlsx"
        p2 = tmp_path / "b.xlsx"
        pl.DataFrame({"x": [1]}).write_excel(str(p1))
        pl.DataFrame({"x": [2]}).write_excel(str(p2))
        out = engine.read_excel([str(p1), str(p2)]).collect()
        paths = sorted(out[FileInfoColumn.FILE_PATH].to_list())
        assert paths == sorted([str(p1), str(p2)])

    def test_read_excel_multi_sheet_has_file_path(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """Multi-sheet result includes __file_path column."""
        path = tmp_path / "multi.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "S1"
        ws1.append(["id"])
        ws1.append([1])
        ws2 = wb.create_sheet("S2")
        ws2.append(["id"])
        ws2.append([2])
        wb.save(str(path))
        result = engine.read_excel(str(path), options={"sheet_id": 0}).collect()
        assert FileInfoColumn.FILE_PATH in result.columns
        assert all(v == str(path) for v in result[FileInfoColumn.FILE_PATH].to_list())

    def test_read_excel_empty_directory_raises(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """Directory with no .xlsx files raises FileNotFoundError."""
        from datacoolie.platforms.local_platform import LocalPlatform
        engine.set_platform(LocalPlatform())
        (tmp_path / "readme.txt").write_text("not excel")
        with pytest.raises(FileNotFoundError, match="No Excel files"):
            engine.read_excel(str(tmp_path))

    def test_read_excel_use_hive_partitioning_stripped(self, engine: PolarsEngine, tmp_path: Path) -> None:
        """use_hive_partitioning is silently removed before calling pl.read_excel."""
        path = tmp_path / "data.xlsx"
        pl.DataFrame({"n": [1]}).write_excel(str(path))
        result = engine.read_excel(str(path), options={"use_hive_partitioning": True})
        assert result.collect().height == 1

    def test_read_path_excel(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = tmp_path / "data.xlsx"
        sample_lf.collect().write_excel(str(path))
        result = engine.read_path(str(path), "excel")
        assert result.collect().height == 3


class TestReadIceberg:
    def test_returns_lazy_frame(self, engine: PolarsEngine, tmp_path: Path) -> None:
        # Polars scan_iceberg is lazy — constructing the frame does not require a catalog.
        # Catalog-based lookup is only required for read_table(..., fmt="iceberg").
        lf = engine.read_iceberg(str(tmp_path))
        assert isinstance(lf, pl.LazyFrame)

    def test_read_table_iceberg_raises_without_catalog(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="iceberg_catalog"):
            engine.read_table("myns.mytable", fmt="iceberg")


class TestReadPath:
    def test_parquet(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = tmp_path / "data.parquet"
        sample_lf.collect().write_parquet(str(path))
        result = engine.read_path(str(path), "parquet")
        assert result.collect().height == 3

    def test_delta(self, engine: PolarsEngine, delta_path: Path) -> None:
        result = engine.read_path(str(delta_path), "delta")
        assert result.collect().height == 3

    def test_unsupported_format(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.read_path("/fake", "xml")


class TestExecuteSQL:
    def test_execute_sql_with_registered_table(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        engine.register_table("people", sample_lf)
        result = engine.execute_sql("SELECT * FROM people WHERE id > 1")
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 2

    def test_execute_sql_aggregation(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        engine.register_table("people", sample_lf)
        result = engine.execute_sql("SELECT COUNT(*) AS cnt FROM people")
        assert result.collect().item() == 3


class TestReadTable:
    def test_raises_without_catalog(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.read_table("some.table")

    def test_unsupported_format(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.read_table("some.table", fmt="avro")


class TestReadDatabase:
    def test_missing_url_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="url"):
            engine.read_database(query="SELECT 1")


# =====================================================================
# Create DataFrame from records
# =====================================================================


class TestCreateDataframe:
    def test_uniform_schema(self, engine: PolarsEngine) -> None:
        records = [{"id": 1, "name": "Alice"}, {"id": 2, "name": "Bob"}]
        lf = engine.create_dataframe(records)
        assert isinstance(lf, pl.LazyFrame)
        df = lf.collect()
        assert df.height == 2
        assert set(df.columns) == {"id", "name"}

    def test_heterogeneous_schema(self, engine: PolarsEngine) -> None:
        """Records with different keys: missing fields become null."""
        records = [
            {"id": 1, "name": "Alice", "score": 9.5},
            {"id": 2, "value": 42},
        ]
        lf = engine.create_dataframe(records)
        df = lf.collect()
        assert df.height == 2
        assert set(df.columns) == {"id", "name", "score", "value"}
        # Missing values are null
        assert df.filter(pl.col("id") == 1)["value"].is_null().all()
        assert df.filter(pl.col("id") == 2)["name"].is_null().all()

    def test_empty_records(self, engine: PolarsEngine) -> None:
        lf = engine.create_dataframe([])
        assert isinstance(lf, pl.LazyFrame)
        assert lf.collect().height == 0

    def test_single_record(self, engine: PolarsEngine) -> None:
        lf = engine.create_dataframe([{"x": 10, "y": "hello"}])
        df = lf.collect()
        assert df.height == 1
        assert df["x"][0] == 10


# =====================================================================
# Write by path
# =====================================================================


class TestWriteToPath:
    def test_write_delta_overwrite(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "out_delta")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="delta")
        result = engine.read_delta(path)
        assert result.collect().height == 3

    def test_write_delta_append(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "out_delta_append")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="delta")
        engine.write_to_path(sample_lf, path, mode="append", fmt="delta")
        result = engine.read_delta(path)
        assert result.collect().height == 6

    def test_write_parquet(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "out_parquet")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="parquet")
        files = list(Path(path).glob("*.parquet"))
        assert len(files) >= 1
        result = engine.read_parquet(str(files[0]))
        assert result.collect().height == 3

    def test_write_csv(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "out_csv")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="csv")
        files = list(Path(path).glob("*.csv"))
        assert len(files) >= 1
        result = engine.read_csv(str(files[0]))
        assert result.collect().height == 3

    def test_write_json(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "out_json")
        Path(path).mkdir(parents=True, exist_ok=True)
        engine._platform = MagicMock()
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="json")
        files = list(Path(path).glob("*.json"))
        assert len(files) >= 1
        result = engine.read_json(str(files[0]))
        assert result.collect().height == 3

    def test_write_jsonl(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "out_jsonl")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="jsonl")
        files = list(Path(path).glob("*.jsonl"))
        assert len(files) >= 1
        result = engine.read_jsonl(str(files[0]))
        assert result.collect().height == 3

    def test_write_parquet_overwrite_clears_folder(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "parquet_ow")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="parquet")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="parquet")
        files = list(Path(path).glob("*.parquet"))
        assert len(files) == 1

    def test_write_parquet_append_adds_file(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "parquet_ap")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="parquet")
        engine.write_to_path(sample_lf, path, mode="append", fmt="parquet")
        files = list(Path(path).glob("*.parquet"))
        assert len(files) == 2

    def test_write_overwrite_file_naming(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "my_table")
        engine.write_to_path(sample_lf, path, mode="overwrite", fmt="parquet")
        files = [f.name for f in Path(path).glob("*.parquet")]
        assert "my_table.parquet" in files

    def test_write_append_file_naming(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        path = str(tmp_path / "my_table")
        engine.write_to_path(sample_lf, path, mode="append", fmt="csv")
        files = [f.name for f in Path(path).glob("*.csv")]
        assert len(files) == 1
        assert files[0].startswith("my_table_") and files[0].endswith(".csv")

    def test_unsupported_format(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.write_to_path(sample_lf, "/fake", mode="overwrite", fmt="xml")


class TestMergeToPath:
    def test_upsert(self, engine: PolarsEngine, delta_path: Path) -> None:
        update_lf = pl.DataFrame(
            {
                "id": [2, 4],
                "name": ["Bobby", "Dan"],
                "date_str": ["2026-01-02", "2026-01-04"],
            }
        ).lazy()
        engine.merge_to_path(update_lf, str(delta_path), merge_keys=["id"])
        result = engine.read_delta(str(delta_path)).collect()
        assert result.height == 4
        bobby = result.filter(pl.col("id") == 2)
        assert bobby["name"][0] == "Bobby"

    def test_non_delta_raises(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="Delta"):
            engine.merge_to_path(sample_lf, "/fake", merge_keys=["id"], fmt="parquet")

    def test_missing_table_raises(self, engine: PolarsEngine, sample_lf: pl.LazyFrame, tmp_path: Path) -> None:
        with pytest.raises(EngineError, match="does not exist"):
            engine.merge_to_path(sample_lf, str(tmp_path / "nonexistent"), merge_keys=["id"])


class TestMergeOverwriteToPath:
    def test_delete_and_reinsert(self, engine: PolarsEngine, delta_path: Path) -> None:
        update_lf = pl.DataFrame(
            {
                "id": [2],
                "name": ["Bobby"],
                "date_str": ["2026-02-02"],
            }
        ).lazy()
        engine.merge_overwrite_to_path(update_lf, str(delta_path), merge_keys=["id"])
        result = engine.read_delta(str(delta_path)).collect()
        assert result.height == 3
        bobby = result.filter(pl.col("id") == 2)
        assert bobby["date_str"][0] == "2026-02-02"


# =====================================================================
# Write by table name (require catalog options)
# =====================================================================


class TestCatalogWrites:
    def test_write_to_table_no_catalog(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.write_to_table(sample_lf, "t", mode="overwrite", fmt="delta")

    def test_merge_to_table_no_catalog(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.merge_to_table(sample_lf, "t", merge_keys=["id"], fmt="delta")

    def test_merge_overwrite_to_table_no_catalog(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.merge_overwrite_to_table(sample_lf, "t", merge_keys=["id"])

    def test_write_to_table_unsupported_format(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.write_to_table(sample_lf, "t", mode="overwrite", fmt="avro")


# =====================================================================
# Transform
# =====================================================================


class TestAddColumn:
    def test_add_literal(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.add_column(sample_lf, "doubled", "id * 2")
        assert isinstance(result, pl.LazyFrame)
        collected = result.collect()
        assert "doubled" in collected.columns
        assert collected["doubled"].to_list() == [2, 4, 6]


class TestDropColumns:
    def test_drop_existing(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.drop_columns(sample_lf, ["name"]).collect()
        assert "name" not in result.columns

    def test_drop_nonexistent(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.drop_columns(sample_lf, ["nonexistent"]).collect()
        assert result.columns == sample_lf.collect().columns


class TestRenameColumn:
    def test_rename(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.rename_column(sample_lf, "name", "full_name").collect()
        assert "full_name" in result.columns
        assert "name" not in result.columns


class TestFilterRows:
    def test_filter(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.filter_rows(sample_lf, "id > 1")
        assert isinstance(result, pl.LazyFrame)
        assert result.collect().height == 2


class TestDeduplicate:
    def test_dedup_desc(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"key": [1, 1, 2], "val": [10, 20, 30]}).lazy()
        result = engine.deduplicate(lf, partition_columns=["key"], order_columns=["val"], order="desc")
        collected = result.collect()
        assert collected.height == 2
        row_key1 = collected.filter(pl.col("key") == 1)
        assert row_key1["val"][0] == 20

    def test_dedup_asc(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"key": [1, 1, 2], "val": [10, 20, 30]}).lazy()
        result = engine.deduplicate(lf, partition_columns=["key"], order_columns=["val"], order="asc")
        collected = result.collect()
        assert collected.height == 2
        row_key1 = collected.filter(pl.col("key") == 1)
        assert row_key1["val"][0] == 10


class TestDeduplicateByRank:
    def test_rank_desc_returns_max(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"key": [1, 1, 1, 2], "val": [10, 10, 20, 30]}).lazy()
        result = engine.deduplicate_by_rank(lf, partition_columns=["key"], order_columns=["val"], order="desc")
        collected = result.collect()
        key1 = collected.filter(pl.col("key") == 1)
        assert key1.height == 1
        assert key1["val"][0] == 20
        # key=2 has only one row, it must be present
        assert collected.filter(pl.col("key") == 2).height == 1

    def test_rank_asc_returns_min(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"key": [1, 1, 2, 2], "val": [5, 10, 3, 7]}).lazy()
        result = engine.deduplicate_by_rank(lf, partition_columns=["key"], order_columns=["val"], order="asc")
        collected = result.collect()
        assert collected.filter(pl.col("key") == 1)["val"][0] == 5
        assert collected.filter(pl.col("key") == 2)["val"][0] == 3

    def test_rank_returns_all_tied_rows(self, engine: PolarsEngine) -> None:
        # key=1 has two rows tied at the max value — both must be returned
        lf = pl.DataFrame({"key": [1, 1, 1], "val": [20, 20, 10]}).lazy()
        result = engine.deduplicate_by_rank(lf, partition_columns=["key"], order_columns=["val"], order="desc")
        collected = result.collect()
        assert collected.height == 2
        assert set(collected["val"].to_list()) == {20}

    def test_rank_multiple_order_columns(self, engine: PolarsEngine) -> None:
        # Lexicographic order: ("date", "amount") desc
        # key=1: (2024-02-01, 100) > (2024-01-01, 200) → row with date 2024-02-01 wins
        # key=2: two rows tied on date, tiebreak by amount → amount=300 wins
        lf = pl.DataFrame({
            "key":    [1,            1,            2,            2           ],
            "date":   ["2024-02-01", "2024-01-01", "2024-02-01", "2024-02-01"],
            "amount": [100,          200,          100,          300         ],
        }).lazy()
        result = engine.deduplicate_by_rank(
            lf,
            partition_columns=["key"],
            order_columns=["date", "amount"],
            order="desc",
        )
        collected = result.collect()
        key1 = collected.filter(pl.col("key") == 1)
        assert key1.height == 1
        assert key1["date"][0] == "2024-02-01"
        assert key1["amount"][0] == 100
        key2 = collected.filter(pl.col("key") == 2)
        assert key2.height == 1
        assert key2["amount"][0] == 300


class TestCastColumn:
    def test_cast_to_string(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.cast_column(sample_lf, "id", "string").collect()
        assert result["id"].dtype == pl.Utf8

    def test_cast_to_long(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"x": [1.0, 2.0, 3.0]}).lazy()
        result = engine.cast_column(lf, "x", "long").collect()
        assert result["x"].dtype == pl.Int64

    def test_cast_to_binary(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"s": ["abc", "def"]}).lazy()
        result = engine.cast_column(lf, "s", "binary").collect()
        assert result["s"].dtype == pl.Binary

    def test_cast_decimal_parameterised(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"amount": [1.5, 2.75, 3.0]}).lazy()
        result = engine.cast_column(lf, "amount", "decimal(10,2)").collect()
        assert result["amount"].dtype == pl.Decimal(precision=10, scale=2)

    def test_cast_date_with_format(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.cast_column(sample_lf, "date_str", "date", fmt="%Y-%m-%d").collect()
        assert result["date_str"].dtype == pl.Date

    def test_unsupported_type_bypasses_cast(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        # Unknown types bypass the cast — column keeps its original dtype
        original_schema = sample_lf.collect_schema()
        result_schema = engine.cast_column(sample_lf, "id", "bizarre_type").collect_schema()
        assert result_schema["id"] == original_schema["id"]


# =====================================================================
# System columns
# =====================================================================


class TestSystemColumns:
    def test_add_system_columns(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.add_system_columns(sample_lf).collect()
        assert SystemColumn.CREATED_AT in result.columns
        assert SystemColumn.UPDATED_AT in result.columns
        assert SystemColumn.UPDATED_BY in result.columns
        assert result[SystemColumn.UPDATED_BY][0] == DEFAULT_AUTHOR

    def test_add_system_columns_custom_author(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.add_system_columns(sample_lf, author="TestBot").collect()
        assert result[SystemColumn.UPDATED_BY][0] == "TestBot"

    def test_remove_system_columns(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with_sys = engine.add_system_columns(sample_lf)
        result = engine.remove_system_columns(with_sys).collect()
        for col in SystemColumn:
            assert col not in result.columns

    def test_add_file_info_columns_with_file_infos(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        """file_infos joined by __file_path injects name and modification_time."""
        path = tmp_path / "data.parquet"
        sample_lf.collect().write_parquet(str(path))
        lf = engine.read_parquet(str(path))  # already has __file_path
        mtime = datetime(2025, 6, 1, tzinfo=timezone.utc)
        fi = FileInfo(name="data.parquet", path=str(path), modification_time=mtime)
        result = engine.add_file_info_columns(lf, file_infos=[fi]).collect()
        assert result[FileInfoColumn.FILE_NAME][0] == "data.parquet"
        # Polars normalises paths to forward slashes internally
        assert result[FileInfoColumn.FILE_PATH][0] == str(path).replace("\\", "/")
        # Compare epoch microseconds to avoid ZoneInfo('UTC') on Windows Store Python.
        expected_us = int(mtime.timestamp() * 1_000_000)
        actual_us = result.select(
            pl.col(FileInfoColumn.FILE_MODIFICATION_TIME).dt.timestamp("us")
        )[FileInfoColumn.FILE_MODIFICATION_TIME][0]
        assert actual_us == expected_us

    def test_add_file_info_columns_no_file_infos(self, engine: PolarsEngine, tmp_path: Path, sample_lf: pl.LazyFrame) -> None:
        """Without file_infos, name is derived from path; mtime is null."""
        path = tmp_path / "data.parquet"
        sample_lf.collect().write_parquet(str(path))
        lf = engine.read_parquet(str(path))  # already has __file_path
        result = engine.add_file_info_columns(lf).collect()
        assert result[FileInfoColumn.FILE_NAME][0] == "data.parquet"
        assert result[FileInfoColumn.FILE_MODIFICATION_TIME][0] is None


# =====================================================================
# Metrics
# =====================================================================


class TestMetrics:
    def test_count_rows(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        assert engine.count_rows(sample_lf) == 3

    def test_is_empty(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        assert not engine.is_empty(sample_lf)
        assert engine.is_empty(pl.DataFrame().lazy())

    def test_get_columns(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        assert engine.get_columns(sample_lf) == ["id", "name", "date_str"]

    def test_get_schema(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        schema = engine.get_schema(sample_lf)
        assert "id" in schema
        assert "name" in schema

    def test_get_max_values(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        maxes = engine.get_max_values(sample_lf, ["id"])
        assert maxes["id"] == 3

    def test_get_count_and_max(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        count, maxes = engine.get_count_and_max_values(sample_lf, ["id"])
        assert count == 3
        assert maxes["id"] == 3


# =====================================================================
# Table operations
# =====================================================================


class TestTableOps:
    def test_exists_by_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        assert engine.table_exists_by_path(str(delta_path)) is True

    def test_not_exists_by_path(self, engine: PolarsEngine, tmp_path: Path) -> None:
        assert engine.table_exists_by_path(str(tmp_path / "no_table")) is False

    def test_exists_by_name_no_catalog(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.table_exists_by_name("some.table")

    def test_history_by_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        history = engine.get_history_by_path(str(delta_path), limit=10)
        assert len(history) >= 1

    def test_history_by_name_no_catalog(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.get_history_by_name("some.table")

    def test_compact_by_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        engine.compact_by_path(str(delta_path))  # should not raise

    def test_compact_by_name_no_catalog(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.compact_by_name("some.table")

    def test_cleanup_by_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        engine.cleanup_by_path(str(delta_path), retention_hours=0)

    def test_cleanup_by_name_no_catalog(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="named Delta"):
            engine.cleanup_by_name("some.table")


# =====================================================================
# Watermark filter
# =====================================================================


class TestWatermarkFilter:
    def test_filter_by_int_watermark(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.apply_watermark_filter(sample_lf, ["id"], {"id": 1})
        assert result.collect().height == 2

    def test_no_watermark_value(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        result = engine.apply_watermark_filter(sample_lf, ["id"], {})
        assert result.collect().height == 3


# =====================================================================
# SQLContext & table registration
# =====================================================================


class TestSQLContext:
    def test_register_table_and_query(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        engine.register_table("t1", sample_lf)
        result = engine.execute_sql("SELECT name FROM t1 WHERE id = 2")
        collected = result.collect()
        assert collected.height == 1
        assert collected["name"][0] == "Bob"

    def test_register_delta_tables(self, tmp_path: Path) -> None:
        from datacoolie.platforms.local_platform import LocalPlatform
        # Create two delta tables in sub-directories
        for name in ("orders", "customers"):
            sub = tmp_path / name
            pl.DataFrame({"col": [1, 2]}).write_delta(str(sub))
        engine = PolarsEngine(platform=LocalPlatform())
        registered = engine.register_delta_tables(str(tmp_path))
        assert "orders" in registered
        assert "customers" in registered
        result = engine.execute_sql("SELECT * FROM orders").collect()
        assert result.height == 2

    def test_register_delta_tables_requires_platform(self, engine: PolarsEngine, tmp_path: Path) -> None:
        with pytest.raises(EngineError, match="requires a platform"):
            engine.register_delta_tables(str(tmp_path))

    def test_sql_context_property(self, engine: PolarsEngine) -> None:
        assert isinstance(engine.sql_context, pl.SQLContext)


# =====================================================================
# Catalog setters
# =====================================================================


class TestCatalogSetters:
    def test_sql_context_injection(self) -> None:
        existing_ctx = pl.SQLContext()
        engine = PolarsEngine(sql_context=existing_ctx)
        assert engine.sql_context is existing_ctx

    def test_default_sql_context_created(self) -> None:
        engine = PolarsEngine()
        assert isinstance(engine.sql_context, pl.SQLContext)

    def test_platform_default_none(self, engine: PolarsEngine) -> None:
        assert engine.platform is None

    def test_set_platform(self, engine: PolarsEngine) -> None:
        from unittest.mock import MagicMock
        from datacoolie.platforms.base import BasePlatform
        mock_platform = MagicMock(spec=BasePlatform)
        engine.set_platform(mock_platform)
        assert engine.platform is mock_platform

    def test_platform_via_init(self) -> None:
        from unittest.mock import MagicMock
        from datacoolie.platforms.base import BasePlatform
        mock_platform = MagicMock(spec=BasePlatform)
        engine = PolarsEngine(platform=mock_platform)
        assert engine.platform is mock_platform

    def test_set_iceberg_catalog(self, engine: PolarsEngine) -> None:
        sentinel = object()
        engine.set_iceberg_catalog(sentinel)
        assert engine._iceberg_catalog is sentinel


# =====================================================================
# Registry integration
# =====================================================================


class TestRegistryIntegration:
    def test_polars_in_engine_registry(self) -> None:
        from datacoolie import engine_registry
        assert engine_registry.is_available("polars")

    def test_create_engine(self) -> None:
        from datacoolie import create_engine
        engine = create_engine("polars")
        assert isinstance(engine, PolarsEngine)


# =====================================================================
# Advanced branch coverage
# =====================================================================


class TestPolarsEngineAdvancedCoverage:
    def test_delta_property_import_error(self, engine: PolarsEngine) -> None:
        with patch.dict("sys.modules", {"deltalake": None}):
            with pytest.raises(EngineError, match="deltalake package is required"):
                _ = engine.delta

    def test_register_delta_tables_list_folders_error_returns_empty(self) -> None:
        mock_platform = MagicMock()
        mock_platform.list_folders.side_effect = RuntimeError("boom")
        engine = PolarsEngine(platform=mock_platform)
        assert engine.register_delta_tables("/root") == []

    def test_register_delta_tables_skips_non_delta_and_scan_errors(self) -> None:
        mock_platform = MagicMock()
        mock_platform.list_folders.return_value = ["/root/a", "/root/b"]
        mock_platform.folder_exists.side_effect = [False, True]
        engine = PolarsEngine(platform=mock_platform)
        with patch("polars.scan_delta", side_effect=RuntimeError("bad scan")):
            assert engine.register_delta_tables("/root") == []

    def test_register_delta_tables_prefix(self, tmp_path: Path) -> None:
        from datacoolie.platforms.local_platform import LocalPlatform
        for name in ("orders", "customers"):
            pl.DataFrame({"col": [1]}).write_delta(str(tmp_path / name))
        engine = PolarsEngine(platform=LocalPlatform())
        registered = engine.register_delta_tables(str(tmp_path), prefix="db1__")
        assert registered == ["db1__customers", "db1__orders"]
        result = engine.execute_sql("SELECT * FROM db1__orders").collect()
        assert result.height == 1

    def test_register_iceberg_tables_requires_input(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="provide namespace"):
            engine.register_iceberg_tables()

    def test_register_iceberg_tables_catalog_discovery(self) -> None:
        catalog = MagicMock()
        catalog.list_tables.return_value = [("ns", "tbl_a"), "tbl_b"]
        catalog.load_table.side_effect = [MagicMock(name="ice_a"), MagicMock(name="ice_b")]
        engine = PolarsEngine(iceberg_catalog=catalog)

        with patch("polars.scan_iceberg", side_effect=[pl.DataFrame({"x": [1]}).lazy(), pl.DataFrame({"y": [2]}).lazy()]):
            registered = engine.register_iceberg_tables(namespace="ns")

        assert registered == ["tbl_a", "tbl_b"]

    def test_register_iceberg_tables_catalog_prefix(self) -> None:
        catalog = MagicMock()
        catalog.list_tables.return_value = [("ns", "sales"), "returns"]
        catalog.load_table.side_effect = [MagicMock(), MagicMock()]
        engine = PolarsEngine(iceberg_catalog=catalog)

        with patch("polars.scan_iceberg", side_effect=[pl.DataFrame({"x": [1]}).lazy(), pl.DataFrame({"y": [2]}).lazy()]):
            registered = engine.register_iceberg_tables(namespace="ns", prefix="cat1__")

        assert registered == ["cat1__sales", "cat1__returns"]

    def test_register_iceberg_tables_base_path_prefix(self) -> None:
        mock_platform = MagicMock()
        mock_platform.list_folders.return_value = ["/root/orders"]
        mock_platform.folder_exists.return_value = True
        engine = PolarsEngine(platform=mock_platform)

        with patch("polars.scan_iceberg", return_value=pl.DataFrame({"id": [1]}).lazy()):
            registered = engine.register_iceberg_tables(base_path="/root", prefix="dw__")

        assert registered == ["dw__orders"]

    def test_register_iceberg_tables_base_path_without_platform_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="requires a platform"):
            engine.register_iceberg_tables(base_path="/root")

    def test_register_iceberg_tables_namespace_without_catalog_and_no_base_path_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="iceberg_catalog is not configured"):
            engine.register_iceberg_tables(namespace="ns")

    def test_register_iceberg_tables_base_path_list_error_returns_empty(self) -> None:
        mock_platform = MagicMock()
        mock_platform.list_folders.side_effect = RuntimeError("boom")
        engine = PolarsEngine(platform=mock_platform)
        assert engine.register_iceberg_tables(base_path="/root") == []

    def test_register_iceberg_tables_base_path_skip_and_register(self) -> None:
        mock_platform = MagicMock()
        mock_platform.list_folders.return_value = ["/root/no_meta", "/root/scan_fail", "/root/ok"]
        mock_platform.folder_exists.side_effect = [False, True, True]
        engine = PolarsEngine(platform=mock_platform)
        def _scan(path: str) -> pl.LazyFrame:
            if path.endswith("scan_fail"):
                raise RuntimeError("bad")
            return pl.DataFrame({"id": [1]}).lazy()

        with patch("polars.scan_iceberg", side_effect=_scan):
            registered = engine.register_iceberg_tables(base_path="/root")
        assert registered == ["ok"]

    def test_read_parquet_maps_basepath_and_storage_options(self) -> None:
        engine = PolarsEngine(storage_options={"token": "x"})
        with patch("polars.scan_parquet", return_value=pl.DataFrame({"a": [1]}).lazy()) as mock_scan:
            _ = engine.read_parquet("/tmp/data.parquet", options={"use_hive_partitioning": "/tmp"})
        kwargs = mock_scan.call_args.kwargs
        assert kwargs["hive_partitioning"] is True
        assert kwargs["storage_options"] == {"token": "x"}
        assert "use_hive_partitioning" not in kwargs

    def test_read_json_list_paths(self, engine: PolarsEngine, tmp_path: Path) -> None:
        p1 = tmp_path / "a.json"
        p2 = tmp_path / "b.json"
        pl.DataFrame({"id": [1]}).write_json(str(p1))
        pl.DataFrame({"id": [2]}).write_json(str(p2))
        out = engine.read_json([str(p1), str(p2)]).collect()
        assert out.height == 2
        assert FileInfoColumn.FILE_PATH in out.columns

    def test_read_database_default_sql_from_table(self, engine: PolarsEngine) -> None:
        with patch("polars.read_database_uri", return_value=pl.DataFrame({"id": [1]})) as mock_read:
            out = engine.read_database(table="users", options={"url": "sqlite:///x.db"})
        assert out.collect().height == 1
        assert mock_read.call_args.args[0] == "SELECT * FROM users"

    def test_read_table_iceberg_with_catalog(self) -> None:
        catalog = MagicMock()
        ice_table = MagicMock()
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)
        with patch("polars.scan_iceberg", return_value=pl.DataFrame({"v": [1]}).lazy()) as mock_scan:
            out = engine.read_table("ns.tbl", fmt="iceberg")
        assert out.collect().height == 1
        assert mock_scan.call_args.args[0] is ice_table

    def test_write_to_path_delta_with_storage_and_partition(self, sample_lf: pl.LazyFrame) -> None:
        engine = PolarsEngine(storage_options={"k": "v"})
        with patch("polars.LazyFrame.sink_delta") as mock_sink_delta:
            engine.write_to_path(sample_lf, "/tmp/delta", mode="full_load", fmt="delta", partition_columns=["id"])
        kwargs = mock_sink_delta.call_args.kwargs
        assert kwargs["mode"] == "overwrite"
        assert kwargs["storage_options"] == {"k": "v"}
        assert kwargs["delta_write_options"] == {"partition_by": ["id"], "schema_mode": "overwrite"}

    def test_merge_overwrite_to_path_non_delta_raises(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="only supports Delta"):
            engine.merge_overwrite_to_path(sample_lf, "/tmp/x", merge_keys=["id"], fmt="parquet")

    def test_merge_overwrite_to_path_missing_delta_table_raises(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="does not exist"):
            engine.merge_overwrite_to_path(sample_lf, "/tmp/nonexistent", merge_keys=["id"])

    def test_write_to_table_iceberg_calls_internal_writer(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with patch.object(engine, "_write_iceberg_table") as mock_write:
            engine.write_to_table(sample_lf, "ns.tbl", mode="append", fmt="iceberg")
        mock_write.assert_called_once()
        assert mock_write.call_args.args[0] is sample_lf

    def test_merge_to_table_iceberg_calls_merge(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with patch.object(engine, "_merge_iceberg_table") as mock_merge:
            engine.merge_to_table(sample_lf, "ns.tbl", merge_keys=["id"], fmt="iceberg")
        mock_merge.assert_called_once_with(sample_lf, "ns.tbl", ["id"], partition_columns=None)

    def test_merge_to_table_unsupported_format_raises(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.merge_to_table(sample_lf, "t", merge_keys=["id"], fmt="parquet")

    def test_merge_overwrite_to_table_iceberg_calls_merge_overwrite(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with patch.object(engine, "_merge_overwrite_iceberg_table") as mock_merge:
            engine.merge_overwrite_to_table(sample_lf, "ns.tbl", merge_keys=["id"], fmt="iceberg")
        mock_merge.assert_called_once_with(sample_lf, "ns.tbl", ["id"], partition_columns=None)

    def test_merge_overwrite_to_table_unsupported_format_raises(self, engine: PolarsEngine, sample_lf: pl.LazyFrame) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.merge_overwrite_to_table(sample_lf, "t", merge_keys=["id"], fmt="parquet")

    def test_apply_watermark_filter_datetime_and_end(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame(
            {
                "id": [1, 2, 3],
                "ts": ["2026-01-01T00:00:00", "2026-01-03T00:00:00", "2026-01-05T00:00:00"],
            }
        ).lazy()
        out = engine.apply_watermark_filter(
            lf,
            ["id", "ts"],
            {"id": 2, "ts": datetime(2026, 1, 2, tzinfo=timezone.utc)},
        ).collect()
        assert out.height >= 1

    def test_cast_timestamp_with_format(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"ts": ["2026-01-01 01:02:03"]}).lazy()
        out = engine.cast_column(lf, "ts", "timestamp", fmt="%Y-%m-%d %H:%M:%S").collect()
        assert str(out["ts"].dtype).startswith("Datetime")

    def test_table_exists_by_path_iceberg_branches_and_unsupported(self, caplog: pytest.LogCaptureFixture) -> None:
        # iceberg_catalog.table_exists does not work with paths — always False without a platform
        catalog = MagicMock()
        engine = PolarsEngine(iceberg_catalog=catalog)
        assert engine.table_exists_by_path("tbl", fmt="iceberg") is False
        catalog.table_exists.assert_not_called()

        assert engine.table_exists_by_path("tbl", fmt="avro") is False
        assert "unsupported format" in caplog.text

    def test_table_exists_by_path_iceberg_without_catalog_returns_false(self) -> None:
        engine = PolarsEngine()
        assert engine.table_exists_by_path("tbl", fmt="iceberg") is False

    def test_table_exists_by_path_delta_with_platform_uses_folder_check(self) -> None:
        mock_platform = MagicMock()
        engine = PolarsEngine(platform=mock_platform)
        mock_platform.folder_exists.return_value = True
        assert engine.table_exists_by_path("/data/tbl") is True
        mock_platform.folder_exists.assert_called_once_with("/data/tbl/_delta_log")

        mock_platform.folder_exists.return_value = False
        assert engine.table_exists_by_path("/data/tbl") is False

    def test_table_exists_by_path_iceberg_with_platform_uses_metadata_check(self) -> None:
        mock_platform = MagicMock()
        engine = PolarsEngine(platform=mock_platform)
        mock_platform.folder_exists.return_value = True
        assert engine.table_exists_by_path("/data/tbl/", fmt="iceberg") is True
        mock_platform.folder_exists.assert_called_once_with("/data/tbl/metadata")

        mock_platform.folder_exists.return_value = False
        assert engine.table_exists_by_path("/data/tbl/", fmt="iceberg") is False

    def test_table_exists_by_name_iceberg_branches(self) -> None:
        engine_no_catalog = PolarsEngine()
        with pytest.raises(EngineError, match="requires iceberg_catalog"):
            engine_no_catalog.table_exists_by_name("tbl", fmt="iceberg")

        catalog = MagicMock()
        catalog.table_exists.return_value = True
        engine = PolarsEngine(iceberg_catalog=catalog)
        assert engine.table_exists_by_name("tbl", fmt="iceberg") is True

        catalog.table_exists.return_value = False
        assert engine.table_exists_by_name("tbl", fmt="iceberg") is False

    def test_table_exists_by_name_unsupported_format_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.table_exists_by_name("tbl", fmt="avro")

    def test_get_history_by_path_branches(self, engine: PolarsEngine, delta_path: Path) -> None:
        assert engine.get_history_by_path(str(delta_path), fmt="parquet") == []

        # Force dt.history failure branch.
        with patch.object(PolarsEngine, "delta", new=property(lambda _self: MagicMock(side_effect=RuntimeError("boom")))):
            assert engine.get_history_by_path(str(delta_path), fmt="delta") == []

        fake_dt = MagicMock()
        fake_dt.history.return_value = [
            {"timestamp": 1735689600000, "version": 1},  # 2025-01-01
            {"timestamp": 1738368000000, "version": 2},  # 2025-02-01
            {"timestamp": 1740787200000, "version": 4},  # 2025-03-01
            {"timestamp": None, "version": 3},
        ]
        with patch.object(PolarsEngine, "delta", new=property(lambda _self: MagicMock(return_value=fake_dt))):
            # No time filter — all entries returned, timestamps normalized to datetime.
            all_entries = engine.get_history_by_path(str(delta_path), limit=10)
            assert all(
                isinstance(e["timestamp"], datetime) or e["timestamp"] is None
                for e in all_entries
            )

            # Filter branches with start/end bounds.
            out = engine.get_history_by_path(
                str(delta_path),
                start_time=datetime(2025, 1, 15, tzinfo=timezone.utc),
                end_time=datetime(2025, 2, 15, tzinfo=timezone.utc),
            )
        assert [e["version"] for e in out] == [2, 3]
        # Filtered entries also have datetime timestamps (None entries pass through).
        assert all(isinstance(e["timestamp"], datetime) or e["timestamp"] is None for e in out)

    def test_get_history_by_name_unsupported_non_delta(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="unsupported format"):
            engine.get_history_by_name("t", fmt="avro")

    def test_get_history_by_name_iceberg_no_catalog_raises(self) -> None:
        engine = PolarsEngine()
        with pytest.raises(EngineError, match="requires iceberg_catalog"):
            engine.get_history_by_name("ns.tbl", fmt="iceberg")

    def test_get_history_by_name_iceberg_load_error_returns_empty(self) -> None:
        catalog = MagicMock()
        catalog.load_table.side_effect = RuntimeError("not found")
        engine = PolarsEngine(iceberg_catalog=catalog)
        assert engine.get_history_by_name("ns.tbl", fmt="iceberg") == []

    def test_get_history_by_name_iceberg_returns_entries(self) -> None:
        snap1 = MagicMock()
        snap1.snapshot_id = 1
        snap1.parent_snapshot_id = None
        snap1.timestamp_ms = 1735689600000  # 2025-01-01
        snap1.manifest_list = "s3://bucket/snap1.avro"
        snap1.summary = {"operation": "append"}
        snap1.dict.return_value = {
            "snapshot_id": 1,
            "parent_snapshot_id": None,
            "timestamp_ms": 1735689600000,
            "manifest_list": "s3://bucket/snap1.avro",
            "summary": {"operation": "append"},
        }

        snap2 = MagicMock()
        snap2.snapshot_id = 2
        snap2.parent_snapshot_id = 1
        snap2.timestamp_ms = 1738368000000  # 2025-02-01
        snap2.manifest_list = "s3://bucket/snap2.avro"
        snap2.summary = {"operation": "overwrite"}
        snap2.dict.return_value = {
            "snapshot_id": 2,
            "parent_snapshot_id": 1,
            "timestamp_ms": 1738368000000,
            "manifest_list": "s3://bucket/snap2.avro",
            "summary": {"operation": "overwrite"},
        }

        ice_table = MagicMock()
        ice_table.snapshots.return_value = [snap1, snap2]

        catalog = MagicMock()
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)

        # No time filter → both snapshots returned, newest first
        result = engine.get_history_by_name("ns.tbl", limit=10, fmt="iceberg")
        assert len(result) == 2
        assert result[0]["snapshot_id"] == 2
        assert result[0]["operation"] == "overwrite"
        assert result[1]["snapshot_id"] == 1
        assert result[1]["operation"] == "append"

    def test_get_history_by_name_iceberg_time_filter(self) -> None:
        snap1 = MagicMock()
        snap1.snapshot_id = 1
        snap1.timestamp_ms = 1735689600000  # 2025-01-01
        snap1.dict.return_value = {
            "snapshot_id": 1,
            "parent_snapshot_id": None,
            "timestamp_ms": 1735689600000,
            "manifest_list": "s3://snap1.avro",
            "summary": {"operation": "append"},
        }

        snap2 = MagicMock()
        snap2.snapshot_id = 2
        snap2.timestamp_ms = 1738368000000  # 2025-02-01
        snap2.dict.return_value = {
            "snapshot_id": 2,
            "parent_snapshot_id": 1,
            "timestamp_ms": 1738368000000,
            "manifest_list": "s3://snap2.avro",
            "summary": {"operation": "overwrite"},
        }

        snap3 = MagicMock()
        snap3.snapshot_id = 3
        snap3.timestamp_ms = 1740787200000  # 2025-03-01
        snap3.dict.return_value = {
            "snapshot_id": 3,
            "parent_snapshot_id": 2,
            "timestamp_ms": 1740787200000,
            "manifest_list": "s3://snap3.avro",
            "summary": {"operation": "append"},
        }

        ice_table = MagicMock()
        ice_table.snapshots.return_value = [snap1, snap2, snap3]

        catalog = MagicMock()
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)

        # start_time=Jan 15 excludes snap1 (ts <= start), end_time=Feb 15 excludes snap3 (ts >= end)
        out = engine.get_history_by_name(
            "ns.tbl",
            limit=10,
            fmt="iceberg",
            start_time=datetime(2025, 1, 15, tzinfo=timezone.utc),
            end_time=datetime(2025, 2, 15, tzinfo=timezone.utc),
        )
        assert [r["snapshot_id"] for r in out] == [2]

    def test_delta_post_commit_checkpoint_and_cleanup(self, tmp_path: Path) -> None:
        """_delta_post_commit creates checkpoint at version %10==0 and cleans metadata."""
        engine = PolarsEngine()
        path = str(tmp_path / "ckpt_table")
        # Create 10 commits (version 0..9) → version 10 triggers checkpoint
        lf = pl.DataFrame({"id": [1]}).lazy()
        for _ in range(10):
            engine._write_delta_path(lf, path, "append", None, {})
        dt = engine.delta(path)
        assert dt.version() == 9
        # Write one more to reach version 10
        engine._write_delta_path(lf, path, "append", None, {})
        # Checkpoint file should exist for version 10
        ckpt_dir = tmp_path / "ckpt_table" / "_delta_log"
        ckpt_files = list(ckpt_dir.glob("*.checkpoint.parquet"))
        assert len(ckpt_files) >= 1

    def test_delta_post_commit_skips_non_checkpoint_version(self, tmp_path: Path) -> None:
        """_delta_post_commit does NOT create a checkpoint when version%10!=0."""
        engine = PolarsEngine()
        path = str(tmp_path / "no_ckpt")
        lf = pl.DataFrame({"id": [1]}).lazy()
        engine._write_delta_path(lf, path, "overwrite", None, {})  # version 0
        engine._write_delta_path(lf, path, "append", None, {})  # version 1
        ckpt_dir = tmp_path / "no_ckpt" / "_delta_log"
        ckpt_files = list(ckpt_dir.glob("*.checkpoint.parquet"))
        assert len(ckpt_files) == 0

    def test_delta_post_commit_error_is_swallowed(self, caplog: pytest.LogCaptureFixture) -> None:
        """_delta_post_commit should not raise on bad path."""
        engine = PolarsEngine()
        # Non-existent path; should log debug and not raise
        engine._delta_post_commit("/tmp/nonexistent_delta_table_xyz")

    def test_compact_cleanup_non_delta_paths_and_name_unsupported(
        self,
        engine: PolarsEngine,
        caplog: pytest.LogCaptureFixture,
    ) -> None:
        engine.compact_by_path("/tmp/x", fmt="parquet")
        engine.cleanup_by_path("/tmp/x", fmt="json")
        assert "only supports Delta" in caplog.text

        # Iceberg compact logs warning (pyiceberg lacks compaction)
        engine.compact_by_name("t", fmt="iceberg")
        assert "does not support compaction" in caplog.text

        # Iceberg cleanup requires catalog
        with pytest.raises(EngineError, match="requires iceberg_catalog"):
            engine.cleanup_by_name("t", fmt="iceberg")

    def test_compact_by_name_iceberg_noop(self, caplog: pytest.LogCaptureFixture) -> None:
        catalog = MagicMock()
        engine = PolarsEngine(iceberg_catalog=catalog)
        engine.compact_by_name("ns.tbl", fmt="iceberg")
        assert "does not support compaction" in caplog.text

    def test_compact_by_name_iceberg_options(self, caplog: pytest.LogCaptureFixture) -> None:
        catalog = MagicMock()
        engine = PolarsEngine(iceberg_catalog=catalog)
        engine.compact_by_name("ns.tbl", fmt="iceberg", options={"rewrite_data_files": False})
        assert "does not support compaction" in caplog.text

    def test_cleanup_by_name_iceberg_expire_snapshots(self) -> None:
        ice_table = MagicMock()
        expire = MagicMock()
        ice_table.maintenance.expire_snapshots.return_value = expire
        expire.older_than.return_value = expire

        # Seed an old snapshot (timestamp well before the 48h cutoff) that is
        # not the current snapshot, so the pre-check allows expire to proceed.
        old_snap = MagicMock()
        old_snap.timestamp_ms = 0
        old_snap.snapshot_id = 111
        ice_table.snapshots.return_value = [old_snap]
        ice_table.metadata.current_snapshot_id = 999

        catalog = MagicMock()
        catalog.load_table.return_value = ice_table

        engine = PolarsEngine(iceberg_catalog=catalog)
        engine.cleanup_by_name("ns.tbl", retention_hours=48, fmt="iceberg")
        catalog.load_table.assert_called_once_with("ns.tbl")
        ice_table.maintenance.expire_snapshots.assert_called_once()
        expire.older_than.assert_called_once()
        expire.commit.assert_called_once()

    def test_cleanup_by_name_iceberg_skip_expire(self, caplog: pytest.LogCaptureFixture) -> None:
        ice_table = MagicMock()
        catalog = MagicMock()
        catalog.load_table.return_value = ice_table

        engine = PolarsEngine(iceberg_catalog=catalog)
        engine.cleanup_by_name(
            "ns.tbl", fmt="iceberg",
            options={"expire_snapshots": False, "remove_orphan_files": False},
        )
        ice_table.maintenance.expire_snapshots.assert_not_called()

    def test_cleanup_by_name_iceberg_remove_orphan_warning(self, caplog: pytest.LogCaptureFixture) -> None:
        ice_table = MagicMock()
        expire = MagicMock()
        ice_table.maintenance.expire_snapshots.return_value = expire
        expire.older_than.return_value = expire
        catalog = MagicMock()
        catalog.load_table.return_value = ice_table

        engine = PolarsEngine(iceberg_catalog=catalog)
        engine.cleanup_by_name("ns.tbl", fmt="iceberg", options={"remove_orphan_files": True})
        assert "does not support remove_orphan_files" in caplog.text

    @patch.object(PolarsEngine, "_reorder_arrow_to_iceberg", side_effect=lambda arrow, ice: arrow)
    def test_write_iceberg_table_branches(self, _mock_reorder: MagicMock) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()

        engine = PolarsEngine()
        with pytest.raises(EngineError, match="requires iceberg_catalog"):
            engine._write_iceberg_table(lf, "tbl", "append")

        ice_table = MagicMock()
        id_field = MagicMock()
        id_field.name = "id"
        ice_table.schema.return_value.fields = [id_field]
        ice_table.spec.return_value.fields = []
        catalog = MagicMock()
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)

        engine._write_iceberg_table(lf, "tbl", "overwrite")
        ice_table.overwrite.assert_called_once()
        ice_table.append.assert_not_called()

        ice_table.reset_mock()
        engine._write_iceberg_table(lf, "tbl", "append")
        ice_table.append.assert_called_once()
        ice_table.overwrite.assert_not_called()

        with pytest.raises(EngineError, match="unsupported Iceberg write mode"):
            engine._write_iceberg_table(lf, "tbl", "merge")


# =====================================================================
# Router override tests — format-aware navigation
# =====================================================================


class TestRouterOverrides:
    """Verify that the PolarsEngine unified routers dispatch based on format:
    Delta → path-based, Iceberg → catalog/name-based."""

    # ---- read ----

    def test_read_delta_routes_to_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        result = engine.read("delta", path=str(delta_path))
        assert isinstance(result, pl.LazyFrame)

    def test_read_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.read("delta", table_name="my_table")

    def test_read_iceberg_with_table_name_routes_to_catalog(self) -> None:
        ice_table = MagicMock()
        catalog = MagicMock()
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)
        with patch("polars.scan_iceberg", return_value=pl.DataFrame({"x": [1]}).lazy()):
            result = engine.read("iceberg", table_name="ns.tbl")
        assert isinstance(result, pl.LazyFrame)

    def test_read_iceberg_with_path_falls_back(self, engine: PolarsEngine) -> None:
        with patch("polars.scan_iceberg", return_value=pl.DataFrame({"x": [1]}).lazy()):
            result = engine.read("iceberg", path="/some/path")
        assert isinstance(result, pl.LazyFrame)

    def test_read_no_args_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="requires table_name or path"):
            engine.read("delta")

    # ---- write ----

    def test_write_delta_routes_to_path(self, engine: PolarsEngine, tmp_path: Path) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        path = str(tmp_path / "write_delta")
        engine.write(lf, path=path, mode="overwrite", fmt="delta")
        assert (tmp_path / "write_delta" / "_delta_log").exists()

    def test_write_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.write(lf, table_name="my_table", mode="overwrite", fmt="delta")

    @patch.object(PolarsEngine, "_reorder_arrow_to_iceberg", side_effect=lambda arrow, ice: arrow)
    def test_write_iceberg_with_table_name_routes_to_catalog(self, _mock_reorder: MagicMock) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        ice_table = MagicMock()
        id_field = MagicMock()
        id_field.name = "id"
        ice_table.schema.return_value.fields = [id_field]
        catalog = MagicMock()
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)
        engine.write(lf, table_name="ns.tbl", mode="overwrite", fmt="iceberg")

    def test_write_iceberg_with_only_path_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="pass table_name instead of path"):
            engine.write(lf, path="/some/path", mode="overwrite", fmt="iceberg")

    def test_write_no_args_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="requires table_name or path"):
            engine.write(lf, mode="overwrite", fmt="delta")

    # ---- merge ----

    def test_merge_delta_routes_to_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        lf = pl.DataFrame({"id": [1], "val": [99]}).lazy()
        engine.merge(lf, path=str(delta_path), merge_keys=["id"], fmt="delta")

    def test_merge_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.merge(lf, table_name="my_table", merge_keys=["id"], fmt="delta")

    def test_merge_iceberg_with_table_name_routes_to_catalog(self) -> None:
        lf = pl.DataFrame({"id": [1], "val": [10]}).lazy()
        catalog = MagicMock()
        engine = PolarsEngine(iceberg_catalog=catalog)
        with patch.object(engine, "merge_to_table") as mock_merge:
            engine.merge(lf, table_name="ns.tbl", merge_keys=["id"], fmt="iceberg")
            mock_merge.assert_called_once()

    def test_merge_iceberg_with_only_path_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="pass table_name instead of path"):
            engine.merge(lf, path="/some/path", merge_keys=["id"], fmt="iceberg")

    def test_merge_no_args_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="requires table_name or path"):
            engine.merge(lf, merge_keys=["id"], fmt="delta")

    # ---- merge_overwrite ----

    def test_merge_overwrite_delta_routes_to_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        lf = pl.DataFrame({"id": [1], "val": [99]}).lazy()
        engine.merge_overwrite(lf, path=str(delta_path), merge_keys=["id"], fmt="delta")

    def test_merge_overwrite_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.merge_overwrite(lf, table_name="my_table", merge_keys=["id"], fmt="delta")

    def test_merge_overwrite_iceberg_with_table_name_routes_to_catalog(self) -> None:
        lf = pl.DataFrame({"id": [1], "val": [10]}).lazy()
        catalog = MagicMock()
        engine = PolarsEngine(iceberg_catalog=catalog)
        with patch.object(engine, "merge_overwrite_to_table") as mock_mo:
            engine.merge_overwrite(lf, table_name="ns.tbl", merge_keys=["id"], fmt="iceberg")
            mock_mo.assert_called_once()

    def test_merge_overwrite_no_args_raises(self, engine: PolarsEngine) -> None:
        lf = pl.DataFrame({"id": [1]}).lazy()
        with pytest.raises(EngineError, match="requires table_name or path"):
            engine.merge_overwrite(lf, merge_keys=["id"], fmt="delta")

    # ---- exists ----

    def test_exists_delta_routes_to_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        assert engine.exists(path=str(delta_path), fmt="delta") is True

    def test_exists_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.exists(table_name="my_table", fmt="delta")

    def test_exists_iceberg_with_table_name_routes_to_catalog(self) -> None:
        catalog = MagicMock()
        catalog.table_exists.return_value = True
        engine = PolarsEngine(iceberg_catalog=catalog)
        assert engine.exists(table_name="ns.tbl", fmt="iceberg") is True

    def test_exists_iceberg_with_path_falls_back(self) -> None:
        mock_platform = MagicMock()
        mock_platform.folder_exists.return_value = True
        engine = PolarsEngine(platform=mock_platform)
        assert engine.exists(path="/some/path", fmt="iceberg") is True

    def test_exists_no_args_returns_false(self, engine: PolarsEngine) -> None:
        assert engine.exists(fmt="delta") is False

    # ---- get_history ----

    def test_get_history_delta_routes_to_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        result = engine.get_history(path=str(delta_path), fmt="delta")
        assert isinstance(result, list)

    def test_get_history_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.get_history(table_name="my_table", fmt="delta")

    def test_get_history_iceberg_with_table_name_routes_to_catalog(self) -> None:
        catalog = MagicMock()
        ice_table = MagicMock()
        ice_table.history.return_value = []
        ice_table.snapshots.return_value = []
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)
        result = engine.get_history(table_name="ns.tbl", fmt="iceberg")
        assert result == []

    def test_get_history_no_args_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="requires table_name or path"):
            engine.get_history(fmt="delta")

    # ---- compact ----

    def test_compact_delta_routes_to_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        engine.compact(path=str(delta_path), fmt="delta")

    def test_compact_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.compact(table_name="my_table", fmt="delta")

    def test_compact_iceberg_with_table_name_routes_to_catalog(self) -> None:
        catalog = MagicMock()
        engine = PolarsEngine(iceberg_catalog=catalog)
        with patch.object(engine, "compact_by_name") as mock_compact:
            engine.compact(table_name="ns.tbl", fmt="iceberg")
            mock_compact.assert_called_once()

    def test_compact_no_args_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="requires table_name or path"):
            engine.compact(fmt="delta")

    # ---- cleanup ----

    def test_cleanup_delta_routes_to_path(self, engine: PolarsEngine, delta_path: Path) -> None:
        engine.cleanup(path=str(delta_path), fmt="delta")

    def test_cleanup_delta_with_table_name_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="pass path instead of table_name"):
            engine.cleanup(table_name="my_table", fmt="delta")

    def test_cleanup_iceberg_with_table_name_routes_to_catalog(self) -> None:
        ice_table = MagicMock()
        expire = MagicMock()
        ice_table.maintenance.expire_snapshots.return_value = expire
        expire.older_than.return_value = expire
        catalog = MagicMock()
        catalog.load_table.return_value = ice_table
        engine = PolarsEngine(iceberg_catalog=catalog)
        engine.cleanup(table_name="ns.tbl", fmt="iceberg")

    def test_cleanup_no_args_raises(self, engine: PolarsEngine) -> None:
        with pytest.raises(EngineError, match="requires table_name or path"):
            engine.cleanup(fmt="delta")


class TestPolarsToHiveType:
    """Test PolarsEngine._polars_type_to_hive with native Polars dtype objects."""

    def test_scalars(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.Int64()) == "BIGINT"
        assert PolarsEngine._polars_type_to_hive(pl.Int32()) == "INT"
        assert PolarsEngine._polars_type_to_hive(pl.Int16()) == "SMALLINT"
        assert PolarsEngine._polars_type_to_hive(pl.Int8()) == "TINYINT"
        assert PolarsEngine._polars_type_to_hive(pl.Float32()) == "FLOAT"
        assert PolarsEngine._polars_type_to_hive(pl.Float64()) == "DOUBLE"
        assert PolarsEngine._polars_type_to_hive(pl.Boolean()) == "BOOLEAN"
        assert PolarsEngine._polars_type_to_hive(pl.String()) == "STRING"
        assert PolarsEngine._polars_type_to_hive(pl.Binary()) == "BINARY"
        assert PolarsEngine._polars_type_to_hive(pl.Date()) == "DATE"

    def test_unsigned_ints(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.UInt8()) == "SMALLINT"
        assert PolarsEngine._polars_type_to_hive(pl.UInt16()) == "INT"
        assert PolarsEngine._polars_type_to_hive(pl.UInt32()) == "BIGINT"
        assert PolarsEngine._polars_type_to_hive(pl.UInt64()) == "BIGINT"

    def test_datetime(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.Datetime("us", "UTC")) == "TIMESTAMP"
        assert PolarsEngine._polars_type_to_hive(pl.Datetime("us")) == "TIMESTAMP"

    def test_decimal(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.Decimal(10, 2)) == "DECIMAL(10,2)"
        assert PolarsEngine._polars_type_to_hive(pl.Decimal(38, 18)) == "DECIMAL(38,18)"

    def test_duration(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.Duration("us")) == "STRING"

    def test_time(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.Time()) == "STRING"

    def test_null(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.Null()) == "STRING"

    def test_list(self) -> None:
        assert PolarsEngine._polars_type_to_hive(pl.List(pl.Int64())) == "ARRAY<BIGINT>"

    def test_nested_list(self) -> None:
        assert PolarsEngine._polars_type_to_hive(
            pl.List(pl.List(pl.String()))
        ) == "ARRAY<ARRAY<STRING>>"

    def test_struct(self) -> None:
        dtype = pl.Struct({"a": pl.Int64(), "b": pl.String()})
        assert PolarsEngine._polars_type_to_hive(dtype) == "STRUCT<a:BIGINT,b:STRING>"

    def test_get_hive_schema(self) -> None:
        lf = pl.DataFrame({
            "id": pl.Series([1, 2], dtype=pl.Int64),
            "name": ["Alice", "Bob"],
            "tags": [[1, 2], [3]],
        }).lazy()
        engine = PolarsEngine()
        result = engine.get_hive_schema(lf)
        assert result == {"id": "BIGINT", "name": "STRING", "tags": "ARRAY<BIGINT>"}
