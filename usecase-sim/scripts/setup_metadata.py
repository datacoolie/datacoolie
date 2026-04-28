"""Fan-out metadata from a single use-cases JSON into multiple targets.

Targets:
    file              : emit YAML + XLSX siblings next to the JSON
    db:<dialect>      : apply schema + seed connections/dataflows/schema_hints
                        (dialects: sqlite, postgresql, mysql, mssql, oracle)
    api-db            : alias for db:postgresql (what pg_api_metadata_server reads)

Examples:
    python usecase-sim/scripts/setup_metadata.py
        # → default: file + db:sqlite + db:postgresql (local workspace)

    python usecase-sim/scripts/setup_metadata.py \
        --json usecase-sim/metadata/file/aws_use_cases.json \
        --workspace-id aws-workspace --targets db:postgresql --truncate

    python usecase-sim/scripts/setup_metadata.py \
        --targets file,db:postgresql,db:mysql,db:mssql,db:oracle --truncate
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from _common import (  # noqa: E402
    DB_URLS,
    LOCAL_USE_CASES_JSON,
    METADATA_DB_DIR,
    setup_logging,
)

logger = setup_logging("setup_metadata")

# Dialect → schema SQL file (kept in metadata/database alongside the DB)
_DIALECT_SCHEMA = {
    "sqlite":     METADATA_DB_DIR / "schema.sql",
    "postgresql": METADATA_DB_DIR / "schema_postgres.sql",
    "mysql":      METADATA_DB_DIR / "schema_mysql.sql",
    "mssql":      METADATA_DB_DIR / "schema_mssql.sql",
    "oracle":     METADATA_DB_DIR / "schema_oracle.sql",
}

_RESERVED_COLS = {"format", "database", "precision"}

ALL_DIALECTS = ["sqlite", "postgresql", "mysql", "mssql", "oracle"]


# ===========================================================================
# DDL helpers (inlined from metadata/database/_shared.py)
# ===========================================================================

def quote_col(name: str, dialect: str) -> str:
    if name not in _RESERVED_COLS:
        return name
    if dialect == "mysql":
        return f"`{name}`"
    if dialect == "mssql":
        return f"[{name}]"
    if dialect == "oracle":
        return name
    return f'"{name}"'


def apply_schema(engine) -> None:
    from sqlalchemy import text  # noqa: PLC0415
    dialect = engine.dialect.name
    schema_path = _DIALECT_SCHEMA.get(dialect)
    if schema_path is None:
        raise ValueError(f"Unsupported dialect: {dialect}")
    if not schema_path.exists():
        raise FileNotFoundError(f"Schema file missing: {schema_path}")
    sql = re.sub(r"--[^\n]*", "", schema_path.read_text(encoding="utf-8"))
    with engine.begin() as conn:
        if dialect == "oracle":
            for block in re.split(r"^\s*/\s*$", sql, flags=re.MULTILINE):
                block = block.strip()
                if block:
                    conn.execute(text(block))
        else:
            for stmt in sql.split(";"):
                stmt = stmt.strip()
                if stmt:
                    conn.execute(text(stmt))
    logger.info("Applied schema: %s (%s)", schema_path.name, dialect)


# ===========================================================================
# Seed helpers (condensed from metadata/database/seed_from_json.py)
# ===========================================================================

def _now() -> datetime:
    return datetime.now(timezone.utc).replace(tzinfo=None)


def _json_str(obj):
    if obj is None:
        return None
    if isinstance(obj, str):
        return obj
    return json.dumps(obj)


def _name_to_uuid(name: str) -> str:
    # Vendored from datacoolie.utils.helpers.name_to_uuid to avoid importing
    # the full datacoolie package.
    import uuid  # noqa: PLC0415
    return str(uuid.uuid5(uuid.UUID("da7ac001-e000-4000-8000-000000000000"), name))


def _connections_sql(dialect: str) -> str:
    q = lambda c: quote_col(c, dialect)  # noqa: E731
    return f"""
    INSERT INTO dc_framework_connections
        (connection_id, workspace_id, name, description,
         connection_type, {q("format")}, catalog, {q("database")},
         configure, secrets_ref, is_active, version,
         created_at, updated_at, created_by, updated_by)
    VALUES
        (:cid, :ws, :name, :desc,
         :ctype, :fmt, :catalog, :db,
         :configure, :secrets_ref, :active, 1,
         :now, :now, 'setup', 'setup')
    """


_DATAFLOWS_SQL = """
INSERT INTO dc_framework_dataflows
    (dataflow_id, workspace_id, name, description,
     stage, group_number, execution_order, processing_mode,
     source_connection_id, source_schema, source_table,
     source_query, source_python_function,
     source_watermark_columns, source_configure,
     transform,
     destination_connection_id, destination_schema, destination_table,
     destination_load_type, destination_merge_keys, destination_configure,
     configure, is_active, version,
     created_at, updated_at, created_by, updated_by)
VALUES
    (:did, :ws, :name, :desc,
     :stage, :gnum, :eorder, :pmode,
     :src_cid, :src_schema, :src_table,
     :src_query, :src_pyfunc,
     :src_wm, :src_conf,
     :transform,
     :dst_cid, :dst_schema, :dst_table,
     :dst_load, :dst_merge, :dst_conf,
     :conf, :active, 1,
     :now, :now, 'setup', 'setup')
"""


def _schema_hints_sql(dialect: str) -> str:
    q = lambda c: quote_col(c, dialect)  # noqa: E731
    return f"""
    INSERT INTO dc_framework_schema_hints
        (schema_hint_id, connection_id, dataflow_id,
         schema_name, table_name, column_name,
         data_type, {q("format")}, {q("precision")}, scale,
         default_value, ordinal_position, is_active,
         created_at, updated_at)
    VALUES
        (:hid, :cid, :dfid,
         :sname, :tname, :cname,
         :dtype, :fmt, :prec, :scale,
         :defval, :ordinal, :active,
         :now, :now)
    """


def _insert_or_skip(conn, sql, params, nested: bool):
    from sqlalchemy import text  # noqa: PLC0415
    from sqlalchemy.exc import IntegrityError  # noqa: PLC0415
    try:
        if nested:
            with conn.begin_nested():
                conn.execute(text(sql), params)
        else:
            conn.execute(text(sql), params)
    except IntegrityError:
        pass


def _truncate_workspace(engine, workspace_id: str) -> None:
    from sqlalchemy import text  # noqa: PLC0415
    with engine.begin() as conn:
        conn.execute(
            text(
                "DELETE FROM dc_framework_schema_hints "
                "WHERE dataflow_id IN "
                "  (SELECT dataflow_id FROM dc_framework_dataflows WHERE workspace_id = :ws) "
                "OR connection_id IN "
                "  (SELECT connection_id FROM dc_framework_connections WHERE workspace_id = :ws)"
            ),
            {"ws": workspace_id},
        )
        conn.execute(text("DELETE FROM dc_framework_dataflows WHERE workspace_id = :ws"),
                     {"ws": workspace_id})
        conn.execute(text("DELETE FROM dc_framework_connections WHERE workspace_id = :ws"),
                     {"ws": workspace_id})
    logger.info("Truncated workspace_id=%s", workspace_id)


def seed_db(connection_string: str, meta: dict, workspace_id: str, truncate: bool) -> None:
    from sqlalchemy import create_engine  # noqa: PLC0415

    engine = create_engine(connection_string)
    dialect = engine.dialect.name
    nested = (dialect == "postgresql")

    apply_schema(engine)
    if truncate:
        _truncate_workspace(engine, workspace_id)

    now = _now()
    conn_sql  = _connections_sql(dialect)
    hints_sql = _schema_hints_sql(dialect)

    connections = meta.get("connections", [])
    with engine.begin() as conn:
        for c in connections:
            cid = c.get("connection_id") or _name_to_uuid(c["name"])
            _insert_or_skip(conn, conn_sql, {
                "cid": cid, "ws": workspace_id,
                "name": c["name"], "desc": c.get("description"),
                "ctype": c["connection_type"],
                "fmt": c.get("format", ""),
                "catalog": c.get("catalog"), "db": c.get("database"),
                "configure": _json_str(c.get("configure")),
                "secrets_ref": _json_str(c.get("secrets_ref")),
                "active": c.get("is_active", True),
                "now": now,
            }, nested)

    dataflows = meta.get("dataflows", [])
    with engine.begin() as conn:
        for d in dataflows:
            did = d.get("dataflow_id") or _name_to_uuid(d["name"])
            src = d.get("source", {}) or {}
            dst = d.get("destination", {}) or {}
            src_cid = (
                d.get("source_connection_id")
                or _name_to_uuid(src.get("connection_name", "") or d.get("source_connection", ""))
            )
            dst_cid = (
                d.get("destination_connection_id")
                or _name_to_uuid(dst.get("connection_name", "") or d.get("destination_connection", ""))
            )

            dst_configure = dst.get("configure") or d.get("destination_configure") or {}
            if isinstance(dst_configure, str):
                try:
                    dst_configure = json.loads(dst_configure)
                except (ValueError, TypeError):
                    dst_configure = {}
            if not isinstance(dst_configure, dict):
                dst_configure = {}
            dst_partitions = dst.get("partition_columns") or d.get("destination_partition_columns")
            if dst_partitions and "partition_columns" not in dst_configure:
                dst_configure = dict(dst_configure)
                dst_configure["partition_columns"] = dst_partitions

            _insert_or_skip(conn, _DATAFLOWS_SQL, {
                "did": did, "ws": workspace_id,
                "name": d["name"], "desc": d.get("description"),
                "stage": d.get("stage"),
                "gnum": d.get("group_number"),
                "eorder": d.get("execution_order"),
                "pmode": d.get("processing_mode"),
                "src_cid": src_cid,
                "src_schema": src.get("schema") or d.get("source_schema"),
                "src_table":  src.get("table")  or d.get("source_table"),
                "src_query":  src.get("query")  or d.get("source_query"),
                "src_pyfunc": src.get("python_function") or d.get("source_python_function"),
                "src_wm":   _json_str(src.get("watermark_columns") or d.get("source_watermark_columns")),
                "src_conf": _json_str(src.get("configure") or d.get("source_configure")),
                "transform": _json_str(d.get("transform")),
                "dst_cid": dst_cid,
                "dst_schema": dst.get("schema") or d.get("destination_schema"),
                "dst_table":  dst.get("table")  or d.get("destination_table"),
                "dst_load":   dst.get("load_type") or d.get("destination_load_type"),
                "dst_merge":  _json_str(dst.get("merge_keys") or d.get("destination_merge_keys")),
                "dst_conf":   _json_str(dst_configure or None),
                "conf": _json_str(d.get("configure")),
                "active": d.get("is_active", True),
                "now": now,
            }, nested)

    schema_hints = meta.get("schema_hints", [])
    with engine.begin() as conn:
        for group in schema_hints:
            cid = group.get("connection_id") or _name_to_uuid(group.get("connection_name", ""))
            columns = group.get("hints", []) or group.get("columns", [])
            for col in columns:
                hint_key = f"{group.get('connection_name', '')}__{group.get('table_name', '')}__{col['column_name']}"
                hid = _name_to_uuid(hint_key)
                _insert_or_skip(conn, hints_sql, {
                    "hid": hid, "cid": cid, "dfid": group.get("dataflow_id"),
                    "sname": group.get("schema_name"),
                    "tname": group.get("table_name"),
                    "cname": col["column_name"],
                    "dtype": col["data_type"],
                    "fmt": col.get("format"),
                    "prec": col.get("precision"),
                    "scale": col.get("scale"),
                    "defval": col.get("default_value"),
                    "ordinal": col.get("ordinal_position"),
                    "active": col.get("is_active", True),
                    "now": now,
                }, nested)

    logger.info("Seeded %s (%s): %d connections, %d dataflows, %d hint groups",
                dialect, workspace_id,
                len(connections), len(dataflows), len(schema_hints))


# ===========================================================================
# File sibling emitters (YAML + XLSX) — condensed from build_metadata_siblings.py
# ===========================================================================

def _json_cell(value):
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _list_to_csv(value):
    if not value:
        return ""
    if isinstance(value, list):
        parts = [json.dumps(i, ensure_ascii=False) if isinstance(i, dict) else str(i) for i in value]
        return ",".join(parts)
    return str(value)


def emit_yaml(data: dict, out_path: Path) -> None:
    try:
        import yaml  # noqa: PLC0415
    except ImportError:
        logger.warning("pyyaml not installed — skipping %s", out_path.name)
        return
    with out_path.open("w", encoding="utf-8") as f:
        yaml.dump(data, f, default_flow_style=False, allow_unicode=True, sort_keys=False)
    logger.info("Wrote %s", out_path)


def emit_xlsx(data: dict, out_path: Path) -> None:
    try:
        import openpyxl  # noqa: PLC0415
    except ImportError:
        logger.warning("openpyxl not installed — skipping %s", out_path.name)
        return
    wb = openpyxl.Workbook()

    ws_c = wb.active
    ws_c.title = "connections"
    ws_c.append(["name", "connection_type", "format", "catalog", "database",
                 "configure", "secrets_ref", "is_active"])
    for c in data.get("connections", []):
        ws_c.append([
            c.get("name", ""), c.get("connection_type", ""), c.get("format", ""),
            c.get("catalog", ""), c.get("database", ""),
            _json_cell(c.get("configure")), _json_cell(c.get("secrets_ref")),
            "" if c.get("is_active", True) else False,
        ])

    ws_d = wb.create_sheet("dataflows")
    ws_d.append([
        "name", "stage", "group_number", "execution_order", "processing_mode",
        "is_active", "configure",
        "source_connection_name", "source_schema_name", "source_table",
        "source_query", "source_watermark_columns", "source_configure",
        "destination_connection_name", "destination_schema_name", "destination_table",
        "destination_load_type", "destination_merge_keys",
        "destination_partition_columns", "destination_configure", "transform",
    ])
    for d in data.get("dataflows", []):
        src = d.get("source", {}) or {}
        dst = d.get("destination", {}) or {}
        tx  = d.get("transform", {}) or {}
        part = dst.get("partition_columns", [])
        part_str = json.dumps(part, ensure_ascii=False) if part and isinstance(part[0], dict) else _list_to_csv(part)
        ws_d.append([
            d.get("name", ""), d.get("stage", ""),
            d.get("group_number", ""), d.get("execution_order", ""),
            d.get("processing_mode", "batch"),
            "" if d.get("is_active", True) else False,
            _json_cell(d.get("configure")),
            src.get("connection_name", ""), src.get("schema_name", ""),
            src.get("table", ""), src.get("query", ""),
            _list_to_csv(src.get("watermark_columns", [])),
            _json_cell(src.get("configure")),
            dst.get("connection_name", ""), dst.get("schema_name", ""),
            dst.get("table", ""), dst.get("load_type", "append"),
            _list_to_csv(dst.get("merge_keys", [])),
            part_str, _json_cell(dst.get("configure")),
            _json_cell(tx) if tx else "",
        ])

    ws_h = wb.create_sheet("schema_hints")
    ws_h.append(["connection_name", "table_name", "schema_name",
                 "column_name", "data_type", "format", "precision", "scale", "is_active"])
    for group in data.get("schema_hints", []):
        cn = group.get("connection_name", "")
        tn = group.get("table_name", "")
        sn = group.get("schema_name", "")
        for h in group.get("hints", []):
            ws_h.append([
                cn, tn, sn,
                h.get("column_name", ""), h.get("data_type", ""),
                h.get("format", ""), h.get("precision", ""), h.get("scale", ""),
                "" if h.get("is_active", True) else False,
            ])

    wb.save(out_path)
    logger.info("Wrote %s", out_path)


# ===========================================================================
# Driver
# ===========================================================================

def _parse_targets(raw: str) -> list[str]:
    items = [t.strip().lower() for t in raw.split(",") if t.strip()]
    valid: list[str] = []
    for t in items:
        if t == "file":
            valid.append("file")
        elif t == "api-db":
            valid.append("db:postgresql")
        elif t.startswith("db:"):
            dialect = t[3:]
            if dialect not in ALL_DIALECTS:
                raise ValueError(f"Unknown dialect: {dialect}. Known: {ALL_DIALECTS}")
            valid.append(t)
        else:
            raise ValueError(f"Unknown target: {t}")
    # dedupe preserving order
    seen: set[str] = set()
    out: list[str] = []
    for t in valid:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def _default_targets() -> str:
    return "file," + ",".join(f"db:{d}" for d in ["sqlite", "postgresql"])


def main() -> int:
    parser = argparse.ArgumentParser(description="Fan-out metadata from JSON to multiple targets")
    parser.add_argument("--json", default=str(LOCAL_USE_CASES_JSON),
                        help="Path to use-cases JSON (default: local_use_cases.json)")
    parser.add_argument("--workspace-id", default="local-workspace")
    parser.add_argument("--targets", default=_default_targets(),
                        help="Comma-separated: file, db:<dialect>, api-db. "
                             "Dialects: " + ", ".join(ALL_DIALECTS))
    parser.add_argument("--truncate", action="store_true",
                        help="Delete existing rows for the workspace before seeding")
    parser.add_argument("--db-url", action="append", default=[],
                        metavar="DIALECT=URL",
                        help="Override a dialect's connection string (repeatable)")
    args = parser.parse_args()

    targets = _parse_targets(args.targets)
    logger.info("Targets: %s", targets)

    # Merge --db-url overrides
    db_urls = dict(DB_URLS)
    for kv in args.db_url:
        if "=" not in kv:
            logger.error("Bad --db-url %r (want DIALECT=URL)", kv)
            return 2
        k, v = kv.split("=", 1)
        if k not in ALL_DIALECTS:
            logger.error("Unknown dialect in --db-url: %s", k)
            return 2
        db_urls[k] = v

    json_path = Path(args.json).resolve()
    if not json_path.exists():
        logger.error("JSON not found: %s", json_path)
        return 2

    with json_path.open("r", encoding="utf-8") as f:
        meta = json.load(f)

    rc = 0
    for target in targets:
        if target == "file":
            logger.info("--- target: file ---")
            emit_yaml(meta, json_path.with_suffix(".yaml"))
            emit_xlsx(meta, json_path.with_suffix(".xlsx"))
        elif target.startswith("db:"):
            dialect = target[3:]
            url = db_urls[dialect]
            logger.info("--- target: db:%s ---", dialect)
            try:
                seed_db(url, meta, args.workspace_id, truncate=args.truncate)
            except Exception as exc:
                logger.warning("db:%s failed: %s", dialect, exc)
                rc = 1
    return rc


if __name__ == "__main__":
    sys.exit(main())
